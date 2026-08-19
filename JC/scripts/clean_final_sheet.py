#!/usr/bin/env python3
"""Clean final_sheet.xlsx → customer_name, city, address, phones, person, outstanding, notes, is_active."""
from __future__ import annotations

import re
import sys
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "final_sheet.xlsx"
OUT = ROOT / "final_sheet_cleaned.xlsx"

sys.path.insert(0, str(ROOT / "scripts"))
from clean_debtors import build_known_cities, extract_parens, extract_phones, split_name  # noqa: E402
from clean_debtors_r2 import CITY_ALIASES, clean_business, clean_city, merge_extra, title_case  # noqa: E402

# ── aliases & junk ─────────────────────────────────────────────────────────────

EXTRA_ALIASES: dict[str, str] = {
    "nogaon": "Naugaon (Dhar)",
    "nogaon dhar": "Naugaon (Dhar)",
    "naugaon": "Naugaon (Dhar)",
    "soyatkala": "Soyat Kala",
    "soyat kala": "Soyat Kala",
    "hosangabad": "Hoshangabad",
    "hoshangabad": "Hoshangabad",
    "narmadapuram": "Narmadapuram",
    "mandsour": "Mandsaur",
    "mansaur": "Mandsaur",
    "mandsaur": "Mandsaur",
    "astha": "Ashta",
    "ashta": "Ashta",
    "amrawati": "Amravati",
    "amravati": "Amravati",
    "biora": "Biaora",
    "biaora": "Biaora",
    "khirkya": "Khirkiya",
    "khirkiya": "Khirkiya",
    "narsullaganj": "Narsulaganj",
    "narsulaganj": "Narsulaganj",
    "narsingarh": "Narsinghgarh",
    "dhamnaud": "Dhamnod",
    "sonkach": "Sonkatch",
    "sonkash": "Sonkatch",
    "sonkatch": "Sonkatch",
    "badwani": "Barwani",
    "barwani": "Barwani",
    "jhalawad": "Jhalawar",
    "jhalawar": "Jhalawar",
    "banswada": "Banswara",
    "baswada": "Banswara",
    "baroda": "Vadodara",
    "mow": "Mhow",
    "mahu": "Mhow",
    "aalot": "Alot",
    "indoree": "Indore",
    "khajuri bazar": "Khajuri Bazar",
    "palyamata": "Palyamata",
    "palya mata": "Palyamata",
    "pidawa": "Pidawa",
    "nisarpur": "Nisarpur",
    "gautampura": "Gautampura",
    "deapalpur": "Depalpur",
    "udaynagar": "Udaynagar",
    "taranaa": "Tarana",
    "chindwara": "Chhindwara",
    "chhindwara": "Chhindwara",
    "ganpur gadi": "Ganpur (Singhana)",
    "ganpur": "Ganpur (Singhana)",
    "dhuliya": "Dhule",
    "dhule": "Dhule",
    "rau": "Rau",
    "dakachia": "Dakachiya",
    "dakachiya": "Dakachiya",
    "vidisa": "Vidisha",
    "vidisha": "Vidisha",
    "banglore": "Bangalore",
    "bangalore": "Bangalore",
    "bhilwada": "Bhilwara",
    "bhilwara": "Bhilwara",
    "nepanagr": "Nepanagar",
    "nepanagar": "Nepanagar",
    "mohali": "Mohali (Manawar)",
    "sadora": "Shadura (Ashoknagar)",
    "shadura": "Shadura (Ashoknagar)",
    "kantaphod": "Kataphord",
    "kataphord": "Kataphord",
    "makshi": "Maksi",
    "maksi": "Maksi",
    "sajapur": "Shajapur",
    "shajapur": "Shajapur",
    "ghatabillod": "Ghatabillod",
    "musakhedi": "Musakhedi",
    "mushakhedi": "Musakhedi",
    "hoshiangabad": "Hoshangabad",
    "kukshi": "Kukshi",
    "sanawad": "Sanawad",
    "khargone": "Khargone",
    "khandwa": "Khandwa",
    "ujjain": "Ujjain",
    "ratlam": "Ratlam",
    "indore": "Indore",
    "dewas": "Dewas",
    "dhar": "Dhar",
    "mhow": "Mhow",
    "sehore": "Sehore",
    "badnawar": "Badnawar",
    "nagda": "Nagda",
    "barnagar": "Barnagar",
    "mahidpur": "Mahidpur",
    "tarana": "Tarana",
    "shujalgpur": "Shujalpur",
    "shujalpur": "Shujalpur",
    "bankhedi": "Bankhedi",
    "harda": "Harda",
    "manawar": "Manawar",
    "maheshwar": "Maheshwar",
    "mangliya": "Mangliya",
    "sawariya": "Sawariya",
    "sawer": "Sawer",
    "dungarpur": "Dungarpur",
    "manpur": "Manpur",
    "thane": "Thane",
    "ajmer": "Ajmer",
    "mumbai": "Mumbai",
    "pachore": "Pachore",
    "pachor": "Pachore",
    "akodiya mandi": "Akodiya Mandi",
    "akodiya-mandi": "Akodiya Mandi",
    "nehrunagar": "Nehru Nagar",
    "satwas": "Satwas",
    "punasa": "Punasa",
    "barwah": "Barwah",
    "pansemal": "Pansemal",
    "rajhot": "Rajhot",
    "nisarpur": "Nisarpur",
    "niwari": "Niwari",
    "shivpuri": "Shivpuri",
    "gwalior": "Gwalior",
    "jabalpur": "Jabalpur",
    "bhopal": "Bhopal",
    "neemuch": "Neemuch",
    "nimach": "Neemuch",
    "jhabua": "Jhabua",
    "alirajpur": "Alirajpur",
    "burhanpur": "Burhanpur",
    "betul": "Betul",
    "raisen": "Raisen",
    "rajgarh": "Rajgarh",
    "seoni": "Seoni",
    "chhindwara": "Chhindwara",
    "kalapipal": "Kalapipal",
    "sardarpur": "Sardarpur",
    "dharampuri": "Dharampuri",
    "mundi": "Mundi",
    "sendhwa": "Sendhwa",
    "piplya": "Piplya",
    "kasrawad": "Kasrawad",
    "badnawar": "Badnawar",
    "thandla": "Thandla",
    "ranapur": "Ranapur",
    "gogawan": "Gogawan",
    "bhanpura": "Bhanpura",
    "manasa": "Manasa",
    "suwasara": "Suwasara",
    "garoth": "Garoth",
    "sitamau": "Sitamau",
    "multai": "Multai",
    "amla": "Amla",
    "anjad": "Anjad",
    "mandleshwar": "Mandleshwar",
    "omkareshwar": "Omkareshwar",
    "sanvad": "Sanawad",
    "sanawad": "Sanawad",
    "raghogarh": "Raghogarh",
    "khilchipur": "Khilchipur",
    "shajapur": "Shajapur",
    "agar": "Agar",
    "agar malwa": "Agar Malwa",
    "pipalrawan": "Pipalrawan",
    "depalpur": "Depalpur",
    "mhow": "Mhow",
    "hatpipliya": "Hatpipliya",
    "badwah": "Barwah",
    "dahhi": "Dahi",
    "dahi": "Dahi",
    "khachrod": "Khachrod",
    "jawad": "Jawad",
    "nimbahera": "Nimbahera",
    "kota": "Kota",
    "baran": "Baran",
    "udaipur": "Udaipur",
    "chittorgarh": "Chittorgarh",
    "surat": "Surat",
    "nagpur": "Nagpur",
    "pune": "Pune",
    "aurangabad": "Aurangabad",
    "hyderabad": "Hyderabad",
    "delhi": "Delhi",
    "kolkata": "Kolkata",
    "chennai": "Chennai",
    "pidawa": "Pidawa",
    "banswara": "Banswara",
    "salumbar": "Salumber",
    "salumber": "Salumber",
    "pratapgarh": "Pratapgarh",
    "mandalgarh": "Mandalgarh",
    "bhilwara": "Bhilwara",
    "ashoknagar": "Ashoknagar",
    "guna": "Guna",
    "vidisha": "Vidisha",
    "khurai": "Khurai",
    "sagar": "Sagar",
    "damoh": "Damoh",
    "tikamgarh": "Tikamgarh",
    "chhatarpur": "Chhatarpur",
    "panna": "Panna",
    "satna": "Satna",
    "rewa": "Rewa",
    "sidhi": "Sidhi",
    "singrauli": "Singrauli",
    "shahdol": "Shahdol",
    "umaria": "Umaria",
    "katni": "Katni",
    "narsinghpur": "Narsinghpur",
    "hoshangabad": "Hoshangabad",
    "itarsi": "Itarsi",
    "pipariya": "Pipariya",
    "sohagpur": "Sohagpur",
    "sonkatch": "Sonkatch",
    "mushakhedi": "Musakhedi",
    "bareli": "Bareli",
    "naugaon": "Naugaon (Dhar)",
    "khategaon": "Khategaon",
    "kannod": "Kannod",
    "tihi": "Tihi",
    "udainagar": "Udainagar",
    "salriya": "Salriya",
    "alot": "Alot",
    "javad": "Jawad",
    "jawra": "Jaora",
    "jaora": "Jaora",
    "dhamnod": "Dhamnod",
    "bareli": "Bareli",
    "namli": "Namli",
    "bid": "Bijawar",
    "bijawar": "Bijawar",
    "singhana": "Singhana",
    "nalkheda": "Nalkheda",
    "pithampur": "Pithampur",
    "sarangi": "Sarangi",
}

JUNK_CITY = {
    "cash", "block", "bloc", "billing", "double", "legacy", "shop", "self",
    "book", "full", "old", "new", "strict", "imli", "g", "j", "r", "su",
    "store", "co", "sales", "retail", "teen", "mata", "mandi", "nadi",
    "balaji", "mane", "gabru", "aaron", "weekly", "strc", "ltd", "pvt",
    "india", "press", "printers", "graphics", "cards", "offset",
    "transport", "trans",
}

CITY_NOISE = re.compile(
    r"""(?ix)
    \b(
        weekly | gst | strc | hprp | jb | teenimli | 3imli |
        carrier\s*pnt | bus\s*only | malya | maa\s*sharda |
        sawariya\s*tr\.? | chk\s*rcpt\s*properly | chk\s*pymnt\s*crfly |
        siddhi\s*vinayak\s*bus\s*madhu | ganpur\s*gadi |
        fanta\s*jila\s*rtlm | jila\s*ratlam | chopati |
        local | sample\s*print | govind
    )\b
    |
    \b\d{2}-\d{2}\b |
    \d+\s*/-
    """
)

PHONEISH = re.compile(r"(?i)\b(?:mo\.?|mob\.?|mb\.?|mobile|phone|tel\.?|telephone)\b")
TALLY_NO = re.compile(r"(?i)\bno\.?\s*\(\d+\)")
PIN = re.compile(r"\b\d{6}\b")
TRANSPORT = re.compile(r"(?i)\b[\w.&/-]+\s+(transport|trans|carrier|parcel)\b")

# State suffixes to strip when cleaning city names
STATE_WORDS = re.compile(
    r"\b(madhya\s+pradesh|rajasthan|maharashtra|gujarat|uttar\s+pradesh|"
    r"mp|raj|up|mh|gj)\b",
    re.I,
)


# ── helpers ────────────────────────────────────────────────────────────────────

def money100(v) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        n = Decimal(str(v).replace(",", "").strip()) * Decimal("100")
    except (InvalidOperation, ValueError):
        return None
    return n.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def fmt_money(v: Decimal | None) -> str:
    return "" if v is None else format(v, "f")


def levenshtein(a: str, b: str) -> int:
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        cur = [i]
        for j, cb in enumerate(b, 1):
            cur.append(min(prev[j] + 1, cur[j - 1] + 1, prev[j - 1] + (ca != cb)))
        prev = cur
    return prev[-1]


def strip_phones_text(text: str) -> str:
    s = str(text or "")
    s = TALLY_NO.sub(" ", s)
    s = PHONEISH.sub(" ", s)
    s = re.sub(r"\b91[6-9]\d{9}\b", " ", s)
    s = re.sub(r"\b[6-9]\d{9}\b", " ", s)
    s = re.sub(r"\b[6-9]\d{2,4}[-\s]\d{5,8}\b", " ", s)
    s = PIN.sub(" ", s)
    s = TRANSPORT.sub(" ", s)
    s = re.sub(r"\bno\.?\s*$", " ", s, flags=re.I)
    s = re.sub(r"[,;/|]+", ",", s)
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"(?:\s*,\s*)+", ", ", s)
    return s.strip(" ,.-")


def person_from_contact(contact: str, tel_phones: set[str]) -> str:
    """Extract person name from Contact Person field."""
    # if it looks like a phone number that's already in telephone, discard
    digits_only = re.sub(r"\D", "", contact or "")
    if len(digits_only) >= 10 and digits_only[-10:] in tel_phones:
        return ""
    leftover = strip_phones_text(contact)
    leftover = re.sub(r"(?i)\b(mo|mob|mb|no|cash|block|billing|retail)\b", " ", leftover)
    leftover = re.sub(r"[^A-Za-z.\s'-]", " ", leftover)
    leftover = re.sub(r"\s+", " ", leftover).strip(" .,-")
    if not leftover or len(leftover) < 3:
        return ""
    if leftover.lower() in JUNK_CITY:
        return ""
    if len(leftover.split()) > 4:
        return ""
    return title_case(leftover)


def canon_city(name: str, known: set[str]) -> str:
    """Canonicalize a city name using aliases and fuzzy matching."""
    raw = re.sub(r"\s+", " ", (name or "")).strip(" -–,")
    if not raw:
        return ""
    key = raw.lower()
    if key in JUNK_CITY:
        return ""
    if key in EXTRA_ALIASES:
        return EXTRA_ALIASES[key]
    if key in CITY_ALIASES:
        return CITY_ALIASES[key]
    # strip state words and retry
    stripped = STATE_WORDS.sub("", raw).strip(" ,-()").strip()
    if stripped and stripped.lower() != key:
        return canon_city(stripped, known)
    # core without parenthetical qualifier
    core = re.sub(r"\([^)]*\)", "", key).strip()
    if core in EXTRA_ALIASES:
        return EXTRA_ALIASES[core]
    if core in CITY_ALIASES:
        return CITY_ALIASES[core]
    # fuzzy match against known (len>=5, distance 1)
    if len(core) >= 5:
        best, best_d = None, 99
        for k in known:
            if abs(len(k) - len(core)) > 1 or min(len(k), len(core)) < 5:
                continue
            d = levenshtein(core, k)
            if d == 1 and d < best_d:
                best, best_d = k, d
        if best:
            if best in EXTRA_ALIASES:
                return EXTRA_ALIASES[best]
            if best in CITY_ALIASES:
                return CITY_ALIASES[best]
            return title_case(best)
    return title_case(raw)


def sanitize_city(city: str, orig: str, known: set[str]) -> tuple[str, str]:
    extra = ""
    raw = re.sub(r"\s+", " ", (city or "")).strip()
    if raw:
        noise = [m.group(0).strip() for m in CITY_NOISE.finditer(raw)]
        if noise:
            extra = merge_extra(*noise)
        raw = CITY_NOISE.sub(" ", raw)
        raw = re.sub(r"\s+", " ", raw).strip(" -–,/()")
        low = re.sub(r"\([^)]*\)", "", raw).strip().lower()
        if low in JUNK_CITY or not raw:
            raw = ""
        else:
            raw = canon_city(raw, known)
            low2 = re.sub(r"\([^)]*\)", "", raw).strip().lower()
            if low2 in JUNK_CITY:
                raw = ""
    if not raw:
        raw = _recover_city(orig, known)
    return raw, extra


# Phrases in party name / address that reliably map to a city
_RECOVER_MAP = {
    "nagda": "Nagda",
    "dharampuri": "Dharampuri",
    "dhar": "Dhar",
    "dewas": "Dewas",
    "indore": "Indore",
    "manawar": "Manawar",
    "mhow": "Mhow",
    "maheshwar": "Maheshwar",
    "badnawar": "Badnawar",
    "mangliya": "Mangliya",
    "sawer": "Sawer",
    "shujalpur": "Shujalpur",
    "thane": "Thane",
    "ajmer": "Ajmer",
    "mumbai": "Mumbai",
    "harda": "Harda",
    "pachore": "Pachore",
    "palyamata": "Palyamata",
    "akodiya mandi": "Akodiya Mandi",
    "akodiya-mandi": "Akodiya Mandi",
    "satwas": "Satwas",
    "punasa": "Punasa",
    "barwah": "Barwah",
    "pansemal": "Pansemal",
    "sehore": "Sehore",
    "rajhot": "Rajhot",
    "ujjain": "Ujjain",
    "ratlam": "Ratlam",
    "khandwa": "Khandwa",
    "pidawa": "Pidawa",
    "khargone": "Khargone",
    "sanawad": "Sanawad",
    "barnagar": "Barnagar",
    "mahidpur": "Mahidpur",
    "kukshi": "Kukshi",
    "nisarpur": "Nisarpur",
    "manpur": "Manpur",
    "shivpuri": "Shivpuri",
    "bankhedi": "Bankhedi",
    "bhopal": "Bhopal",
    "gwalior": "Gwalior",
    "jabalpur": "Jabalpur",
    "neemuch": "Neemuch",
    "jhabua": "Jhabua",
    "alirajpur": "Alirajpur",
    "burhanpur": "Burhanpur",
    "betul": "Betul",
    "ashoknagar": "Ashoknagar",
    "vidisha": "Vidisha",
    "narsinghpur": "Narsinghpur",
    "itarsi": "Itarsi",
    "kalapipal": "Kalapipal",
    "sardarpur": "Sardarpur",
    "mundi": "Mundi",
    "sendhwa": "Sendhwa",
    "kasrawad": "Kasrawad",
    "thandla": "Thandla",
    "ranapur": "Ranapur",
    "manasa": "Manasa",
    "garoth": "Garoth",
    "sitamau": "Sitamau",
    "anjad": "Anjad",
    "mandleshwar": "Mandleshwar",
    "omkareshwar": "Omkareshwar",
    "agar malwa": "Agar Malwa",
    "agar": "Agar",
    "pipalrawan": "Pipalrawan",
    "depalpur": "Depalpur",
    "hatpipliya": "Hatpipliya",
    "dahi": "Dahi",
    "khachrod": "Khachrod",
    "jawad": "Jawad",
    "alot": "Alot",
    "jaora": "Jaora",
    "dhamnod": "Dhamnod",
    "bareli": "Bareli",
    "kannod": "Kannod",
    "khategaon": "Khategaon",
    "sonkatch": "Sonkatch",
    "double chowki": "Indore",
    "namli": "Namli",
    "bijawar": "Bijawar",
    "badwani": "Barwani",
    "barwani": "Barwani",
    "naugaon": "Naugaon (Dhar)",
    "nogaon": "Naugaon (Dhar)",
    "pipariya": "Pipariya",
    "itarsi": "Itarsi",
    "hoshangabad": "Hoshangabad",
}


def _recover_city(blob: str, known: set[str]) -> str:
    text = " " + re.sub(r"[^A-Za-z\s()-]", " ", (blob or "").lower()) + " "
    for key in sorted(_RECOVER_MAP, key=len, reverse=True):
        if re.search(rf"\b{re.escape(key)}\b", text):
            return canon_city(_RECOVER_MAP[key], known)
    return ""


def city_from_address(addr: str, known: set[str]) -> str:
    text = strip_phones_text(addr)
    if not text:
        return ""
    blob = " " + re.sub(r"[^A-Za-z\s()-]", " ", text.lower()) + " "
    pool = sorted(
        set(known) | set(EXTRA_ALIASES) | set(CITY_ALIASES),
        key=len, reverse=True,
    )
    for key in pool:
        if len(key) < 3 or key in JUNK_CITY:
            continue
        if re.search(rf"\b{re.escape(key)}\b", blob):
            return canon_city(key, known)
    chunks = [c.strip() for c in re.split(r"[,]", text) if c.strip()]
    for ch in reversed(chunks):
        cand = re.sub(r"[^A-Za-z\s()-]", " ", ch).strip()
        if not cand:
            continue
        low = cand.lower()
        if low in JUNK_CITY:
            continue
        if low in known or low in EXTRA_ALIASES or low in CITY_ALIASES:
            return canon_city(cand, known)
        words = cand.split()
        if len(words) == 1 and len(words[0]) >= 4 and words[0].isalpha():
            return canon_city(words[0], known)
    return ""


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ── main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    wb = load_workbook(SRC, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # find data start (row after header row with Debit/Credit)
    start = 12
    for i, r in enumerate(rows):
        vals = [str(c).strip().lower() if c is not None else "" for c in (r or [])]
        if "debit" in vals and "credit" in vals:
            start = i + 1
            break

    # collect raw party rows
    raw_rows = []
    names = []
    for r in rows[start:]:
        if not r or not r[0] or not str(r[0]).strip():
            continue
        name = str(r[0]).strip()
        if name.lower() in {"grand total", "total"}:
            continue
        names.append(name)
        raw_rows.append(r)

    # build known cities from party names + aliases
    known: set[str] = build_known_cities(names)
    known |= {k.lower() for k in EXTRA_ALIASES}
    known |= {k.lower() for k in CITY_ALIASES}

    cleaned = []
    for r in raw_rows:
        raw = str(r[0]).strip()
        addr = str(r[1]).strip() if r[1] else ""
        contact = str(r[2]).strip() if r[2] else ""
        tel = str(r[3]).strip() if r[3] else ""
        bills_raw = r[4]
        receipt_raw = r[5]
        debit_raw = r[6]
        credit_raw = r[7]
        status_raw = str(r[8]).strip() if r[8] else ""
        is_active = status_raw.upper() != "INACTIVE PARTIES"

        # ── phones ──
        p, s, phone_notes = extract_phones(tel, addr)
        if not p:
            p2, s2, n2 = extract_phones(contact, None)
            if p2:
                p, s = p2, s2 or s
                phone_notes = merge_extra(phone_notes, n2)

        # phones already in tel → used to decide if contact is just a phone
        tel_phone_set: set[str] = set()
        if p:
            tel_phone_set.add(p)
        if s:
            tel_phone_set.add(s)

        # ── person ──
        person = person_from_contact(contact, tel_phone_set)

        # ── city from party name ──
        biz0, city0, extra0 = split_name(raw, known)
        biz, city_from_biz, extra_from_biz = clean_business(biz0, city0, known)
        city_name, extra_from_city_field = clean_city(city0 or city_from_biz, known)
        extra = merge_extra(extra0, extra_from_biz, extra_from_city_field)

        # ── city from address ──
        city_addr = city_from_address(addr, known)

        city = ""
        if city_name:
            city = canon_city(city_name, known)
        if city_addr:
            ca = canon_city(city_addr, known)
            if not city:
                city = ca
            elif ca and ca.lower() != city.lower():
                extra = merge_extra(extra, f"addr_city:{ca}")

        city, city_extra = sanitize_city(city, f"{raw} {addr}", known)
        extra = merge_extra(extra, city_extra)

        # ── business name fallback ──
        if not biz:
            stub, tags = extract_parens(raw)
            extra = merge_extra(extra, *tags)
            if city:
                stub = re.sub(rf"(?i)\b{re.escape(city.split('(')[0].strip())}\b", " ", stub)
            stub = re.sub(r"(?i)\b(cash|block|billing|retail)\b", " ", stub)
            stub = re.sub(r"\s+", " ", stub).strip(" -–,")
            if stub:
                biz = stub

        # ── address (cleaned, minus phones, minus city) ──
        clean_addr = strip_phones_text(addr)
        if city:
            city_core = city.split("(")[0].strip()
            clean_addr = re.sub(rf"(?i)\b{re.escape(city_core)}\b", " ", clean_addr)
        clean_addr = re.sub(r"(?i)\b(madhya pradesh|rajasthan|maharashtra|gujarat|m\.p\.?)\b", " ", clean_addr)
        clean_addr = re.sub(r"\s+", " ", clean_addr).strip(" ,.-")
        # if all that's left is noise (only state/pin etc.), clear it
        if len(clean_addr) < 4:
            clean_addr = ""

        # ── financials ──
        debit = money100(debit_raw)
        credit = money100(credit_raw)
        d = debit if debit is not None else Decimal("0.00")
        c = credit if credit is not None else Decimal("0.00")
        outstanding = (d - c).quantize(Decimal("0.01"))

        # ── notes ──
        bills = int(bills_raw) if bills_raw is not None and str(bills_raw).strip() else 0
        receipt = int(receipt_raw) if receipt_raw is not None and str(receipt_raw).strip() else 0
        counts_note = ""
        if bills or receipt:
            counts_note = f"Bills:{bills} Receipt:{receipt}"
        money_note = f"Tally D:{fmt_money(debit) or '0'} C:{fmt_money(credit) or '0'}"
        notes = merge_extra(extra, phone_notes, counts_note, money_note)

        # ── review reasons (flag only genuinely problematic rows) ──
        reasons = []
        if not biz:
            reasons.append("no_business")
        if not city:
            reasons.append("no_city")

        cleaned.append({
            "customer_name": title_case(biz) if biz else "",
            "city": city,
            "address": clean_addr,
            "primary_phone": p,
            "secondary_phone": s,
            "person_name": person,
            "outstanding": fmt_money(outstanding),
            "is_active": is_active,
            "notes": notes,
            "original_name": raw,
            "review": " | ".join(reasons),
        })

    # ── write output ──
    out = Workbook()
    headers = [
        "customer_name", "city", "address",
        "primary_phone", "secondary_phone", "person_name",
        "outstanding", "is_active", "notes", "original_name",
    ]
    ws1 = out.active
    ws1.title = "Cleaned"
    ws1.append(headers)

    ws2 = out.create_sheet("Needs Review")
    ws2.append(headers + ["review_reason"])

    ws3 = out.create_sheet("Inactive")
    ws3.append(headers)

    n_ok = n_rev = n_inactive = 0
    for rec in cleaned:
        row = [rec[h] for h in headers]
        if not rec["is_active"]:
            ws3.append(row)
            n_inactive += 1
        elif rec["review"]:
            ws2.append(row + [rec["review"]])
            n_rev += 1
        else:
            ws1.append(row)
            n_ok += 1

    col_widths = (36, 22, 32, 14, 14, 20, 14, 10, 54, 44)
    for sheet in (ws1, ws2, ws3):
        style_header(sheet)
        for col, w in zip("ABCDEFGHIJ", col_widths):
            sheet.column_dimensions[col].width = w

    # widen review_reason col in ws2
    ws2.column_dimensions["K"].width = 28

    out.save(OUT)

    unique_cities = sorted({r["city"] for r in cleaned if r["city"]})
    no_city = sum(1 for r in cleaned if not r["city"])
    no_phone = sum(1 for r in cleaned if not r["primary_phone"])
    print(f"total rows   : {len(cleaned)}")
    print(f"active-clean : {n_ok}")
    print(f"needs review : {n_rev}")
    print(f"inactive     : {n_inactive}")
    print(f"unique cities: {len(unique_cities)}")
    print(f"no city      : {no_city}")
    print(f"no phone     : {no_phone}")
    outstanding_sum = sum(Decimal(r["outstanding"]) for r in cleaned)
    print(f"outstanding sum: ₹{outstanding_sum:,.2f}")
    print(f"wrote → {OUT}")
    if no_city:
        print("\nno_city originals (active only):")
        for r in cleaned:
            if not r["city"] and r["is_active"]:
                print(f"  {r['original_name']}")


if __name__ == "__main__":
    main()
