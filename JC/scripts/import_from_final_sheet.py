#!/usr/bin/env python3
"""
Wipe customer data only, then import from final_sheet_cleaned.xlsx.
Sheets loaded: Cleaned (active), Needs Review (active), Inactive.
Skips rows with no valid phone.
Sets credit_limit = 0 for all.
"""
from __future__ import annotations

import os
import re
import sys
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "backend"
SRC = ROOT / "final_sheet_cleaned.xlsx"
OUT = ROOT / "final_sheet_import_result.xlsx"
AS_ON = date(2026, 6, 30)

sys.path.insert(0, str(BACKEND))
os.chdir(BACKEND)

from dotenv import load_dotenv  # noqa: E402
load_dotenv(BACKEND / ".env")

from sqlalchemy import func, text  # noqa: E402
from openpyxl import Workbook, load_workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from app.db.session import SessionLocal  # noqa: E402
from app.models.city import City  # noqa: E402
from app.models.customer import Customer  # noqa: E402
from app.models.accounts_receivable import CustomerArAccount, ArLedgerEntry  # noqa: E402
from app.services.passwords import hash_password  # noqa: E402


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ── wipe customer data only ────────────────────────────────────────────────────

def wipe_customers(db) -> dict:
    stmts = [
        ("jc_customer_return_lines", "DELETE FROM jc_customer_return_lines"),
        ("jc_customer_returns", "DELETE FROM jc_customer_returns"),
        ("jc_ar_ledger_entries", "DELETE FROM jc_ar_ledger_entries"),
        ("jc_customer_bill_lines", "DELETE FROM jc_customer_bill_lines"),
        ("jc_customer_bills", "DELETE FROM jc_customer_bills"),
        ("jc_customer_order_lines", "DELETE FROM jc_customer_order_lines"),
        ("jc_customer_order_placements", "DELETE FROM jc_customer_order_placements"),
        ("jc_customer_open_lines", "DELETE FROM jc_customer_open_lines"),
        ("jc_customer_orders", "DELETE FROM jc_customer_orders"),
        ("jc_customer_ar_accounts", "DELETE FROM jc_customer_ar_accounts"),
        ("jc_customers", "DELETE FROM jc_customers"),
    ]
    counts = {}
    for name, sql in stmts:
        try:
            res = db.execute(text(sql))
            counts[name] = res.rowcount
        except Exception as e:
            db.rollback()
            raise RuntimeError(f"wipe failed on {name}: {e}") from e
    db.commit()
    return counts


# ── helpers ────────────────────────────────────────────────────────────────────

def normalize_phone(v) -> str | None:
    if v is None:
        return None
    d = re.sub(r"\D+", "", str(v))
    if len(d) == 11 and d.startswith("0"):
        d = d[1:]
    if len(d) == 12 and d.startswith("91"):
        d = d[2:]
    return d if len(d) == 10 and d[0] in "6789" else None


def parse_outstanding(v) -> Decimal | None:
    if v is None or str(v).strip() == "":
        return None
    try:
        return Decimal(str(v).replace(",", "").strip()).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def get_or_create_city(db, name: str | None, cache: dict[str, int]) -> int | None:
    if not name or not str(name).strip():
        return None
    clean = re.sub(r"\s+", " ", str(name).strip())
    key = clean.lower()
    if key in cache:
        return cache[key]
    existing = (
        db.query(City)
        .filter(City.deleted_at.is_(None), func.lower(City.name) == key)
        .one_or_none()
    )
    if existing:
        cache[key] = existing.id
        return existing.id
    row = City(name=clean, route_id=None, is_active=True)
    db.add(row)
    db.flush()
    cache[key] = row.id
    return row.id


def set_opening_balance_signed(db, *, customer_id: int, amount: Decimal, as_on: date) -> None:
    """Upsert signed opening_balance AR entry. Positive = they owe us. Negative = we owe them."""
    # ensure AR account exists
    ar = db.query(CustomerArAccount).filter(CustomerArAccount.customer_id == customer_id).first()
    if not ar:
        ar = CustomerArAccount(customer_id=customer_id, is_open=True)
        db.add(ar)
        db.flush()
    # remove any existing opening_balance entries
    db.query(ArLedgerEntry).filter(
        ArLedgerEntry.customer_id == customer_id,
        ArLedgerEntry.entry_type == "opening_balance",
    ).delete(synchronize_session=False)
    db.flush()
    if amount == Decimal("0.00"):
        return
    direction = "debit" if amount > 0 else "credit"
    entry = ArLedgerEntry(
        customer_id=customer_id,
        entry_type="opening_balance",
        amount=amount,  # signed: positive = owed to us
        description=f"Opening balance as on {as_on.isoformat()} [{direction} ₹{abs(amount):,.2f}]",
        value_date=as_on,
        created_by_type="admin",
        created_by_id=None,
        created_by_name="import",
        created_at=datetime(as_on.year, as_on.month, as_on.day, tzinfo=timezone.utc),
    )
    db.add(entry)
    db.flush()


# ── load sheets ────────────────────────────────────────────────────────────────

def load_rows() -> list[dict]:
    wb = load_workbook(SRC, data_only=True)
    out = []
    # sheet name → is_active
    sheet_map = {"Cleaned": True, "Needs Review": True, "Inactive": False}
    for sheet_name, is_active in sheet_map.items():
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: sheet '{sheet_name}' not found, skipping")
            continue
        ws = wb[sheet_name]
        header = [str(h or "").strip() for h in next(ws.iter_rows(max_row=1, values_only=True))]
        idx = {h: i for i, h in enumerate(header)}

        def g(r, col, default=None):
            i = idx.get(col)
            return r[i] if i is not None and i < len(r) else default

        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or not g(r, "customer_name"):
                continue
            out.append({
                "customer_name": str(g(r, "customer_name", "")).strip(),
                "city": str(g(r, "city", "") or "").strip() or None,
                "address": str(g(r, "address", "") or "").strip() or None,
                "primary_phone": g(r, "primary_phone"),
                "secondary_phone": g(r, "secondary_phone"),
                "person_name": str(g(r, "person_name", "") or "").strip() or None,
                "outstanding": g(r, "outstanding"),
                "notes": str(g(r, "notes", "") or "").strip() or None,
                "original_name": str(g(r, "original_name", "") or "").strip() or None,
                "is_active": is_active,
            })
    wb.close()
    return out


# ── main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    db = SessionLocal()
    try:
        print("Wiping customer data…")
        counts = wipe_customers(db)
        for k, v in counts.items():
            print(f"  {k}: deleted {v}")

        print(f"Loading {SRC}…")
        rows = load_rows()
        print(f"  sheet rows: {len(rows)}")

        city_cache: dict[str, int] = {}
        used_phones: set[str] = set()
        created: list[dict] = []
        skipped: list[dict] = []

        for row in rows:
            biz = row["customer_name"]
            phone = normalize_phone(row["primary_phone"])
            secondary = normalize_phone(row["secondary_phone"])
            if secondary == phone:
                secondary = None

            if not phone:
                skipped.append({**row, "reason": "no_phone"})
                continue
            if not biz:
                skipped.append({**row, "reason": "no_name"})
                continue
            if phone in used_phones:
                skipped.append({**row, "reason": "duplicate_phone"})
                continue

            city_id = get_or_create_city(db, row["city"], city_cache)
            route_id = None
            if city_id:
                city_row = db.get(City, city_id)
                route_id = city_row.route_id if city_row else None

            outstanding = parse_outstanding(row["outstanding"])

            cust = Customer(
                business_name=biz[:500],
                person_name=(row["person_name"] or "")[:500] or None,
                phone=phone,
                password_hash=hash_password(phone[-4:]),
                secondary_phone=secondary,
                address=(row["address"] or "")[:500] or None,
                additional_details=row["notes"],
                city_id=city_id,
                route_id=route_id,
                credit_limit=Decimal("0.00"),
                credit_override=False,
                is_active=row["is_active"],
                # deleted_at stays None for both active and inactive
            )
            db.add(cust)
            db.flush()

            if outstanding is not None:
                set_opening_balance_signed(db, customer_id=cust.id, amount=outstanding, as_on=AS_ON)

            used_phones.add(phone)
            created.append({**row, "customer_id": cust.id, "phone": phone})

            if len(created) % 100 == 0:
                db.commit()
                print(f"  imported {len(created)}…")

        db.commit()
        print(f"\nImported : {len(created)}")
        print(f"Skipped  : {len(skipped)}")

        # active/inactive counts
        n_active = sum(1 for r in created if r["is_active"])
        n_inactive = sum(1 for r in created if not r["is_active"])
        print(f"  active  : {n_active}")
        print(f"  inactive: {n_inactive}")

        # ── result workbook ──
        out_wb = Workbook()
        ws_ok = out_wb.active
        ws_ok.title = "Imported"
        ok_cols = ["customer_id", "is_active", "customer_name", "city", "primary_phone",
                   "secondary_phone", "person_name", "outstanding", "original_name"]
        ws_ok.append(ok_cols)
        for r in created:
            ws_ok.append([r.get(c) for c in ok_cols])

        ws_skip = out_wb.create_sheet("Skipped")
        skip_cols = ["reason", "customer_name", "city", "primary_phone", "outstanding", "original_name"]
        ws_skip.append(skip_cols)
        for r in skipped:
            ws_skip.append([r.get(c) for c in skip_cols])

        for ws in (ws_ok, ws_skip):
            style_header(ws)
        for col, w in zip("ABCDEFGHI", [12, 10, 30, 20, 14, 14, 18, 12, 40]):
            ws_ok.column_dimensions[col].width = w
        for col, w in zip("ABCDEF", [20, 30, 20, 14, 12, 40]):
            ws_skip.column_dimensions[col].width = w

        out_wb.save(OUT)
        print(f"Report  → {OUT}")

        # sanity
        n_total = db.query(Customer).count()
        n_act = db.query(Customer).filter(Customer.is_active.is_(True), Customer.deleted_at.is_(None)).count()
        n_inact = db.query(Customer).filter(Customer.is_active.is_(False), Customer.deleted_at.is_(None)).count()
        print(f"\nDB verify — total:{n_total}  active:{n_act}  inactive:{n_inact}")

    finally:
        db.close()


if __name__ == "__main__":
    main()
