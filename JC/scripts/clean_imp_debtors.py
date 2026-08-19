#!/usr/bin/env python3
"""Clean IMP DEBTORS.xlsx → business, city, person, phones, notes, opening."""
from __future__ import annotations

import os
import re
import sys
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "IMP DEBTORS.XLSX"
OUT = ROOT / "imp_debtors_cleaned.xlsx"
BACKEND = ROOT / "backend"

sys.path.insert(0, str(ROOT / "scripts"))
from clean_debtors import (  # noqa: E402
    build_known_cities,
    extract_parens,
    extract_phones,
    split_name,
)
from clean_debtors_r2 import (  # noqa: E402
    CITY_ALIASES,
    clean_business,
    clean_city,
    merge_extra,
    title_case,
)

EXTRA_ALIASES = {
    "nogaon": "Naugaon (Dhar)",
    "nogaon dhar": "Naugaon (Dhar)",
    "naugaon": "Naugaon (Dhar)",
    "soyatkala": "Soyat Kala",
    "soyat kala": "Soyat Kala",
    "hosangabad": "Hoshangabad",
    "hoshangabad": "Hoshangabad",
    "mandsour": "Mandsaur",
    "mansaur": "Mandsaur",
    "mandsaur": "Mandsaur",
    "astha": "Ashta",
    "ashta": "Ashta",
    "amrawati": "Amravati",
    "amravati": "Amravati",
    "biora": "Biaora",
    "biaora": "Biaora",
    "khirkya": "Khirkiya",
    "khirkiya": "Khirkiya",
    "narsullaganj": "Narsulaganj",
    "narsulaganj": "Narsulaganj",
    "narsingarh": "Narsinghgarh",
    "dhamnaud": "Dhamnod",
    "sonkach": "Sonkatch",
    "sonkash": "Sonkatch",
    "sonkatch": "Sonkatch",
    "badwani": "Barwani",
    "barwani": "Barwani",
    "jhalawad": "Jhalawar",
    "jhalawar": "Jhalawar",
    "banswada": "Banswara",
    "baswada": "Banswara",
    "baroda": "Vadodara",
    "mow": "Mhow",
    "mahu": "Mhow",
    "aalot": "Alot",
    "indoree": "Indore",
    "khajuri bazar": "Khajuri Bazar",
    "palyamata": "Palyamata",
    "pidawa": "Pidawa",
    "nisarpur": "Nisarpur",
    "gautampura": "Gautampura",
    "deapalpur": "Depalpur",
    "udaynagar": "Udaynagar",
    "taranaa": "Tarana",
    "chindwara": "Chhindwara",
    "chhindwara": "Chhindwara",
    "ganpur gadi": "Ganpur (Singhana)",
    "ganpur": "Ganpur (Singhana)",
    "dhuliya": "Dhule",
    "dhule": "Dhule",
    "rau": "Rau",
    "dakachia": "Dakachiya",
    "dakachiya": "Dakachiya",
    "vidisa": "Vidisha",
    "vidisha": "Vidisha",
    "banglore": "Bangalore",
    "bangalore": "Bangalore",
    "bhilwada": "Bhilwara",
    "bhilwara": "Bhilwara",
    "nepanagr": "Nepanagar",
    "nepanagar": "Nepanagar",
    "mohali": "Mohali (Manawar)",
    "sadora": "Shadura (Ashoknagar)",
    "shadura": "Shadura (Ashoknagar)",
    "kantaphod": "Kataphord",
    "kataphord": "Kataphord",
    "makshi": "Maksi",
    "maksi": "Maksi",
    "sajapur": "Shajapur",
    "shajapur": "Shajapur",
    "ghatabillod": "Ghatabillod",
    "musakhedi": "Musakhedi",
    "mushakhedi": "Musakhedi",
    "narmadapuram": "Narmadapuram",
    "hoshiangabad": "Hoshangabad",
}

JUNK_CITY = {
    "cash", "block", "bloc", "billing", "double", "legacy", "shop", "self",
    "book", "full", "old", "new", "strict", "imli", "g", "j", "r", "su",
    "store", "co", "sales", "retail",     "teen", "mata", "mandi", "nadi",
    "balaji", "mane", "gabru", "aaron", "weekly", "strc",
}

CITY_NOISE = re.compile(
    r"""(?ix)
    \b(
        weekly | gst | strc | hprp | jb | teenimli | 3imli |
        carrier\s*pnt | bus\s*only | malya | maa\s*sharda |
        sawariya\s*tr\.? | chk\s*rcpt\s*properly | chk\s*pymnt\s*crfly |
        siddhi\s*vinayak\s*bus\s*madhu | ganpur\s*gadi |
        fanta\s*jila\s*rtlm | jila\s*ratlam | chopati |
        local | sample\s*print | govind
    )\b
    |
    \b\d{2}-\d{2}\b |
    \d+\s*/-
    """
)

RECOVER = {
    "nagda": "Nagda",
    "dharampuri": "Dharampuri",
    "dhar": "Dhar",
    "dewas": "Dewas",
    "indore": "Indore",
    "manawar": "Manawar",
    "mhow": "Mhow",
    "maheshwar": "Maheshwar",
    "badnawar": "Badnawar",
    "mangliya": "Mangliya",
    "sawariya": "Sawariya",
    "dungarpur": "Dungarpur",
    "manpur": "Manpur",
    "sawer": "Sawer",
    "shujalpur": "Shujalpur",
    "thane": "Thane",
    "ajmer": "Ajmer",
    "mumbai": "Mumbai",
    "harda": "Harda",
    "pachor": "Pachore",
    "pachore": "Pachore",
    "palya mata": "Palyamata",
    "palyamata": "Palyamata",
    "akodiya mandi": "Akodiya Mandi",
    "akodiya-mandi": "Akodiya Mandi",
    "akhodiya mandi": "Akhodiya Mandi",
    "chote": "Chote",
    "nehrunagar": "Nehru Nagar",
    "satwas": "Satwas",
    "punasa": "Punasa",
    "barwah": "Barwah",
    "pathachopati": "Dhar",
    "balaji tower": "Indore",
    "store indore": "Indore",
    "co. indore": "Indore",
    "local indore": "Indore",
    "double chowki": "Indore",
    "rain mhow": "Mhow",
    "nadi road": "Pansemal",
    "pansemal": "Pansemal",
    "sehore": "Sehore",
    "rajhot": "Rajhot",
    "ujjain": "Ujjain",
    "ratlam": "Ratlam",
    "khandwa": "Khandwa",
    "pidawa": "Pidawa",
    "khargone": "Khargone",
}

PHONEISH = re.compile(
    r"(?i)\b(?:mo\.?|mob\.?|mb\.?|mobile|phone|tel\.?|telephone)\b"
)
TALLY_NO = re.compile(r"(?i)\bno\.?\s*\(\d+\)")
PIN = re.compile(r"\b\d{6}\b")
TRANSPORT = re.compile(
    r"(?i)\b[\w.&/-]+\s+(transport|trans|carrier|parcel)\b"
)


def money100(v) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        n = Decimal(str(v).replace(",", "").strip()) * Decimal("100")
    except (InvalidOperation, ValueError):
        return None
    return n.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def fmt_money(v: Decimal | None) -> str:
    if v is None:
        return ""
    return format(v, "f")


def levenshtein(a: str, b: str) -> int:
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        cur = [i]
        for j, cb in enumerate(b, 1):
            cur.append(min(prev[j] + 1, cur[j - 1] + 1, prev[j - 1] + (ca != cb)))
        prev = cur
    return prev[-1]


def load_db_cities() -> dict[str, str]:
    """lower-name → canonical DB name, skipping junk tags."""
    try:
        from dotenv import load_dotenv

        load_dotenv(BACKEND / ".env")
        os.chdir(BACKEND)
        sys.path.insert(0, str(BACKEND))
        from app.db.session import SessionLocal, init_db
        from app.models.city import City

        init_db()
        db = SessionLocal()
        try:
            rows = db.query(City).filter(City.deleted_at.is_(None)).all()
        finally:
            db.close()
        out = {}
        for c in rows:
            key = re.sub(r"\s+", " ", c.name).strip().lower()
            if not key or key in JUNK_CITY:
                continue
            core = re.sub(r"\([^)]*\)", "", key).strip()
            if core in JUNK_CITY:
                continue
            display = title_case(c.name) if c.name.isupper() and len(c.name) > 2 else c.name
            out.setdefault(key, display)
            if core and core not in out:
                out[core] = display
        return out
    except Exception:
        return {}


def canon_city(name: str, db_map: dict[str, str], known: set[str]) -> str:
    raw = re.sub(r"\s+", " ", (name or "")).strip(" -–,")
    if not raw:
        return ""
    key = raw.lower()
    if key in JUNK_CITY:
        return ""
    if key in EXTRA_ALIASES:
        raw = EXTRA_ALIASES[key]
        key = raw.lower()
    elif key in CITY_ALIASES:
        raw = CITY_ALIASES[key]
        key = raw.lower()
    if key in db_map:
        return db_map[key]
    core = re.sub(r"\([^)]*\)", "", key).strip()
    if core in EXTRA_ALIASES:
        return EXTRA_ALIASES[core]
    if core in CITY_ALIASES:
        return CITY_ALIASES[core]
    if core in db_map:
        return db_map[core]
    # close typo vs known / db (len>=5, distance 1)
    if len(core) >= 5:
        best = None
        best_d = 99
        pool = set(db_map) | known | set(EXTRA_ALIASES) | set(CITY_ALIASES)
        for k in pool:
            if abs(len(k) - len(core)) > 1:
                continue
            if min(len(k), len(core)) < 5:
                continue
            d = levenshtein(core, k)
            if d == 1 and d < best_d:
                best, best_d = k, d
        if best:
            if best in EXTRA_ALIASES:
                return EXTRA_ALIASES[best]
            if best in CITY_ALIASES:
                return CITY_ALIASES[best]
            if best in db_map:
                return db_map[best]
            return title_case(best)
    return title_case(raw)


def strip_phones_text(text: str) -> str:
    s = str(text or "")
    s = TALLY_NO.sub(" ", s)
    s = PHONEISH.sub(" ", s)
    s = re.sub(r"\b91[6-9]\d{9}\b", " ", s)
    s = re.sub(r"\b[6-9]\d{9}\b", " ", s)
    s = re.sub(r"\b[6-9]\d{2,4}[-\s]\d{5,8}\b", " ", s)
    s = PIN.sub(" ", s)
    s = TRANSPORT.sub(" ", s)
    s = re.sub(r"\bno\.?\s*$", " ", s, flags=re.I)
    s = re.sub(r"[,;/|]+", ",", s)
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"(?:\s*,\s*)+", ", ", s)
    return s.strip(" ,.-")


def person_from_contact(contact: str) -> str:
    leftover = strip_phones_text(contact)
    leftover = re.sub(r"(?i)\b(mo|mob|mb|no|cash|block|billing|retail)\b", " ", leftover)
    leftover = re.sub(r"[^A-Za-z.\s'-]", " ", leftover)
    leftover = re.sub(r"\s+", " ", leftover).strip(" .,-")
    if not leftover or len(leftover) < 3:
        return ""
    if leftover.lower() in JUNK_CITY:
        return ""
    if len(leftover.split()) > 4:
        return ""
    return title_case(leftover)


def recover_city(blob: str, db_map: dict[str, str], known: set[str]) -> str:
    text = " " + re.sub(r"[^A-Za-z\s()-]", " ", (blob or "").lower()) + " "
    for key in sorted(RECOVER, key=len, reverse=True):
        if re.search(rf"\b{re.escape(key)}\b", text):
            return canon_city(RECOVER[key], db_map, known)
    return ""


def sanitize_city(city: str, original: str, db_map: dict[str, str], known: set[str]) -> tuple[str, str]:
    extra = ""
    raw = re.sub(r"\s+", " ", (city or "")).strip()
    if raw:
        noise = [m.group(0).strip() for m in CITY_NOISE.finditer(raw)]
        if noise:
            extra = merge_extra(*noise)
        raw = CITY_NOISE.sub(" ", raw)
        raw = re.sub(r"\s+", " ", raw).strip(" -–,/")
        raw = re.sub(r"\(\s*\)", "", raw).strip()
        low = re.sub(r"\([^)]*\)", "", raw).strip().lower()
        if low in JUNK_CITY or not raw:
            raw = ""
        else:
            raw = canon_city(raw, db_map, known)
            low2 = re.sub(r"\([^)]*\)", "", raw).strip().lower()
            if low2 in JUNK_CITY:
                raw = ""
    if not raw:
        raw = recover_city(original, db_map, known)
    return raw, extra


def city_from_address(addr: str, known: set[str], db_map: dict[str, str]) -> str:
    text = strip_phones_text(addr)
    if not text:
        return ""
    # scan comma chunks right-to-left (city often last)
    chunks = [c.strip() for c in re.split(r"[,]", text) if c.strip()]
    pool = sorted(
        set(known) | set(db_map) | set(EXTRA_ALIASES) | set(CITY_ALIASES),
        key=len,
        reverse=True,
    )
    blob = " " + re.sub(r"[^A-Za-z\s()-]", " ", text.lower()) + " "
    for key in pool:
        if len(key) < 3 or key in JUNK_CITY:
            continue
        if re.search(rf"\b{re.escape(key)}\b", blob):
            return canon_city(key, db_map, known)
    for ch in reversed(chunks):
        cand = re.sub(r"[^A-Za-z\s()-]", " ", ch).strip()
        if not cand:
            continue
        low = cand.lower()
        if low in JUNK_CITY:
            continue
        if low in known or low in db_map or low in EXTRA_ALIASES or low in CITY_ALIASES:
            return canon_city(cand, db_map, known)
        words = cand.split()
        if len(words) == 1 and len(words[0]) >= 4 and words[0].isalpha():
            return canon_city(words[0], db_map, known)
    return ""


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def main() -> None:
    wb = load_workbook(SRC, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    start = 12
    for i, r in enumerate(rows):
        vals = [str(c).strip().lower() if c is not None else "" for c in (r or [])]
        if "debit" in vals and "credit" in vals:
            start = i + 1
            break

    raw_rows = []
    names = []
    for r in rows[start:]:
        if not r or not r[0] or not str(r[0]).strip():
            continue
        raw = str(r[0]).strip()
        if raw.lower() in {"grand total", "total"}:
            continue
        names.append(raw)
        raw_rows.append(r)

    known = build_known_cities(names)
    known |= {k.lower() for k in EXTRA_ALIASES}
    known |= {k.lower() for k in CITY_ALIASES}
    db_map = load_db_cities()
    known |= set(db_map)

    cleaned = []
    for r in raw_rows:
        raw = str(r[0]).strip()
        addr = str(r[1]).strip() if r[1] else ""
        contact = str(r[2]).strip() if r[2] else ""
        tel = str(r[3]).strip() if r[3] else ""
        debit_raw, credit_raw = r[6], r[7]

        biz0, city0, extra0 = split_name(raw, known)
        biz, city_from_biz, extra_from_biz = clean_business(biz0, city0, known)
        city_name, extra_from_city = clean_city(city0 or city_from_biz, known)
        extra = merge_extra(extra0, extra_from_biz, extra_from_city)

        city_addr = city_from_address(addr, known, db_map)
        city = ""
        if city_name:
            city = canon_city(city_name, db_map, known)
        if city_addr:
            ca = canon_city(city_addr, db_map, known)
            if not city:
                city = ca
            elif ca and ca.lower() != city.lower():
                extra = merge_extra(extra, f"addr_city: {ca}")

        city, city_extra = sanitize_city(city, f"{raw} {addr}", db_map, known)
        extra = merge_extra(extra, city_extra)

        if not biz:
            stub, tags = extract_parens(raw)
            extra = merge_extra(extra, *tags)
            if city:
                stub = re.sub(rf"(?i)\b{re.escape(city.split('(')[0].strip())}\b", " ", stub)
            stub = re.sub(r"(?i)\b(cash|block|billing|retail)\b", " ", stub)
            stub = re.sub(r"\s+", " ", stub).strip(" -–,")
            if stub:
                biz = stub

        person = person_from_contact(contact)
        p, s, phone_notes = extract_phones(tel, addr)
        if not p:
            p2, s2, n2 = extract_phones(contact, None)
            if p2:
                p, s, phone_notes = p2, s2 or s, merge_extra(phone_notes, n2)
        leftover_addr = strip_phones_text(addr)
        leftover_addr = re.sub(r"(?i)\b" + re.escape(city.split("(")[0].strip()) + r"\b", " ", leftover_addr) if city else leftover_addr
        leftover_addr = re.sub(r"\s+", " ", leftover_addr).strip(" ,.-")
        if leftover_addr and len(leftover_addr) > 8:
            extra = merge_extra(extra, leftover_addr)

        debit = money100(debit_raw)
        credit = money100(credit_raw)
        d = debit if debit is not None else Decimal("0.00")
        c = credit if credit is not None else Decimal("0.00")
        opening = (d - c).quantize(Decimal("0.01"))

        money_note = f"debit {fmt_money(d)} | credit {fmt_money(c)}"
        notes = merge_extra(extra, phone_notes, money_note)

        reasons = []
        if not biz:
            reasons.append("no_business")
        if not city:
            reasons.append("no_city")
        if not p:
            reasons.append("no_phone")
        if debit is None and credit is None:
            reasons.append("no_amount")

        cleaned.append({
            "original_name": raw,
            "business_name": title_case(biz) if biz else "",
            "person_name": person,
            "city": city,
            "primary_phone": p,
            "secondary_phone": s,
            "notes": notes,
            "debit": fmt_money(debit) if debit is not None else "",
            "credit": fmt_money(credit) if credit is not None else "",
            "opening_balance": fmt_money(opening),
            "review": " | ".join(reasons),
        })

    out = Workbook()
    ws1 = out.active
    ws1.title = "Cleaned"
    headers = [
        "business_name", "person_name", "city",
        "primary_phone", "secondary_phone", "notes",
        "debit", "credit", "opening_balance", "original_name",
    ]
    ws1.append(headers)
    ws2 = out.create_sheet("Needs Review")
    ws2.append(headers + ["review_reason"])

    n_ok = n_rev = 0
    for rec in cleaned:
        row = [rec[h] for h in headers]
        if rec["review"]:
            ws2.append(row + [rec["review"]])
            n_rev += 1
        else:
            ws1.append(row)
            n_ok += 1

    for sheet in (ws1, ws2):
        style_header(sheet)
        for col, w in zip("ABCDEFGHIJ", (36, 22, 24, 14, 14, 48, 12, 12, 16, 44)):
            sheet.column_dimensions[col].width = w

    out.save(OUT)
    cities = sorted({r["city"] for r in cleaned if r["city"]})
    no_city = sum(1 for r in cleaned if not r["city"])
    no_phone = sum(1 for r in cleaned if not r["primary_phone"])
    print(f"rows {len(cleaned)}  clean {n_ok}  review {n_rev}")
    print(f"unique cities {len(cities)}  no_city {no_city}  no_phone {no_phone}")
    print(f"opening sum {sum(Decimal(r['opening_balance']) for r in cleaned)}")
    print(f"wrote {OUT}")
    print("no_city originals:")
    for r in cleaned:
        if not r["city"]:
            print(" ", r["original_name"])


if __name__ == "__main__":
    main()
