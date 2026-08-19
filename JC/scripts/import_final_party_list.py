#!/usr/bin/env python3
"""
Import FINAL PARTY LIST FOR SOFTWARE.XLSX into the database.
- Wipes all existing customer data first.
- Cleans address, contact person, telephone.
- Sorts by bills DESC, receipts DESC, name ASC → assigns party_number.
- Creates missing cities.
- Sets opening balances via AR ledger.
"""
import os, re, sys, hashlib, random, string
from decimal import Decimal
from datetime import date

import openpyxl
import psycopg2
import psycopg2.extras

DB_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://postgres:Bk0gNohhSeELKCB7@db.jovuafnpuhogmngjmpzd.supabase.co:5432/postgres",
)

XLSX_PATH = os.path.join(os.path.dirname(__file__), "..", "FINAL PARTY LIST FOR SOFTWARE.XLSX")


# ─── Phone cleaning ────────────────────────────────────────────────────────────
_PHONE_PAT = re.compile(r"\b[6-9]\d{9}\b")  # 10-digit mobile
_LANDLINE_PAT = re.compile(r"\b0\d{2,4}[-\s]\d{6,8}\b")


def extract_phones(raw: str) -> list[str]:
    """Return up to 2 phone/mobile numbers from raw telephone string."""
    raw = raw.strip()
    # Remove "MO." / "Mo." / "Mobile" prefixes
    raw = re.sub(r"\b(MO|Mo|Mobile|mob|MOBILE)\b\.?\s*", "", raw, flags=re.IGNORECASE)
    found = _PHONE_PAT.findall(raw)
    seen = []
    for p in found:
        if p not in seen:
            seen.append(p)
    return seen[:2]  # max 2


def clean_address(raw: str) -> str:
    """Remove phone numbers and 'No. (XX)' from address."""
    if not raw:
        return ""
    # Remove "No. (XX)" patterns
    raw = re.sub(r",?\s*No\.\s*\(\d+\)", "", raw)
    # Remove mobile numbers
    raw = re.sub(r"\bMO\.?\s*[6-9]\d{9}\b", "", raw, flags=re.IGNORECASE)
    raw = re.sub(r"\b[6-9]\d{9}\b", "", raw)
    # Remove standalone "MO." if leftover
    raw = re.sub(r"\bMO\.?\s*,?", "", raw, flags=re.IGNORECASE)
    # Collapse multiple commas/spaces
    raw = re.sub(r",\s*,+", ",", raw)
    raw = re.sub(r"\s{2,}", " ", raw)
    raw = raw.strip(" ,")
    return raw


def clean_person(raw: str) -> str:
    """Return person name, or empty if it looks like a phone number."""
    if not raw:
        return ""
    raw = raw.strip()
    if _PHONE_PAT.search(raw) and len(re.sub(r"\D", "", raw)) >= 10:
        return ""
    # Pure digits → not a name
    if re.fullmatch(r"[\d\s\-+().]+", raw):
        return ""
    return raw


def normalize_phone(raw: str) -> str:
    digits = re.sub(r"\D", "", raw)
    return digits[-10:] if len(digits) >= 10 else ""


# ─── Load workbook ─────────────────────────────────────────────────────────────
def load_rows():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb.active
    rows = []
    for r in range(12, ws.max_row + 1):
        sr_no = ws.cell(r, 1).value
        if not sr_no:
            continue
        def cell(c):
            return str(ws.cell(r, c).value or "").strip()

        party_name = cell(3)  # Clean party name (col C)
        if not party_name:
            continue

        city_raw = cell(4)
        marker_1 = cell(5)
        marker_2 = cell(6)
        other_notes = cell(7)
        address_raw = cell(8)
        contact_raw = cell(9)
        telephone_raw = cell(10)
        total_bills_raw = ws.cell(r, 11).value
        total_receipts_raw = ws.cell(r, 12).value
        debit_raw = ws.cell(r, 13).value
        credit_raw = ws.cell(r, 14).value
        inactive_raw = cell(15)
        cash_credit = cell(16).upper() or "CREDIT"
        limit_raw = ws.cell(r, 17).value

        # Clean fields
        address = clean_address(address_raw)
        person = clean_person(contact_raw)
        phones = extract_phones(telephone_raw)
        primary_phone = phones[0] if phones else ""
        secondary_phone = phones[1] if len(phones) > 1 else ""

        # Opening balance: (debit - credit) * 100 = rupees owed by customer
        debit = float(debit_raw or 0)
        credit = float(credit_raw or 0)
        opening_balance = round((debit - credit) * 100, 2)

        total_bills = int(float(total_bills_raw)) if total_bills_raw else 0
        total_receipts = int(float(total_receipts_raw)) if total_receipts_raw else 0

        is_active = "INACTIVE" not in inactive_raw.upper()
        try:
            credit_limit = float(limit_raw) if limit_raw and cash_credit == "CREDIT" else 0.0
        except (ValueError, TypeError):
            credit_limit = 0.0

        # Notes: combine other_notes + additional info
        note_parts = []
        if other_notes:
            note_parts.append(other_notes)
        note_parts.append(f"Bills: {total_bills}, Receipts: {total_receipts}, Debit: {debit}, Credit: {credit}")
        notes = " | ".join(note_parts)

        rows.append({
            "party_name": party_name,
            "city": city_raw.upper() if city_raw else "",
            "marker_1": marker_1 or None,
            "marker_2": marker_2 or None,
            "address": address or None,
            "person": person or None,
            "primary_phone": primary_phone,
            "secondary_phone": secondary_phone or None,
            "opening_balance": opening_balance,
            "is_active": is_active,
            "payment_type": cash_credit or "CREDIT",
            "credit_limit": credit_limit,
            "notes": notes,
            "total_bills": total_bills,
            "total_receipts": total_receipts,
        })
    return rows


def hash_password(plain: str) -> str:
    return hashlib.sha256(plain.encode()).hexdigest()


def make_placeholder_phone() -> str:
    suffix = "".join(random.choices(string.digits, k=7))
    return f"000{suffix}"


def main():
    print(f"Loading {XLSX_PATH} …")
    rows = load_rows()
    print(f"  Loaded {len(rows)} parties")

    # Sort: bills DESC, receipts DESC, name ASC
    rows.sort(key=lambda r: (-r["total_bills"], -r["total_receipts"], r["party_name"]))

    # Assign party_number
    for i, r in enumerate(rows, start=1):
        r["party_number"] = i

    print("Connecting to DB …")
    conn = psycopg2.connect(DB_URL)
    conn.autocommit = False
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    print("Collecting cities …")
    cur.execute("SELECT id, name FROM jc_cities WHERE deleted_at IS NULL")
    existing_cities = {row["name"].upper(): row["id"] for row in cur.fetchall()}

    # Collect all unique cities needed
    needed_cities = set(r["city"] for r in rows if r["city"] and r["city"] not in existing_cities)
    city_id_map = dict(existing_cities)
    for city_name in sorted(needed_cities):
        cur.execute(
            "INSERT INTO jc_cities (name) VALUES (%s) ON CONFLICT DO NOTHING RETURNING id",
            (city_name.title(),)
        )
        result = cur.fetchone()
        if result:
            city_id_map[city_name] = result["id"]
        else:
            cur.execute("SELECT id FROM jc_cities WHERE UPPER(name) = %s", (city_name,))
            r2 = cur.fetchone()
            if r2:
                city_id_map[city_name] = r2["id"]
    print(f"  {len(needed_cities)} new cities created")

    # ── WIPE existing customer data ────────────────────────────────────────────
    print("Wiping existing customer data …")
    # Order: AR ledger, orders, bills, then customers
    for tbl in [
        "jc_ar_ledger_entries",
        "jc_customer_bill_lines", "jc_customer_bills",
        "jc_customer_open_lines",
        "jc_customer_order_lines", "jc_customer_order_placements", "jc_customer_orders",
        "jc_customer_ar_accounts",
        "jc_customers",
    ]:
        try:
            cur.execute(f"DELETE FROM {tbl}")
            print(f"  Deleted from {tbl}: {cur.rowcount} rows")
        except Exception as e:
            conn.rollback()
            print(f"  WARNING: could not delete from {tbl}: {e}")
            conn.autocommit = False

    conn.commit()

    # ── Insert customers ───────────────────────────────────────────────────────
    print(f"Inserting {len(rows)} customers …")
    used_phones: set[str] = set()
    imported = 0
    skipped = 0

    for r in rows:
        phone = r["primary_phone"]
        if not phone or len(phone) != 10:
            phone = make_placeholder_phone()
            while phone in used_phones:
                phone = make_placeholder_phone()

        if phone in used_phones:
            phone = make_placeholder_phone()
            while phone in used_phones:
                phone = make_placeholder_phone()
        used_phones.add(phone)

        city_name = r["city"]
        city_id = city_id_map.get(city_name) if city_name else None

        plain_pw = phone[-4:] if len(phone) >= 4 else "0000"
        pw_hash = hash_password(plain_pw)

        credit_limit = r["credit_limit"]
        # CASH parties → credit_limit = 0 (track-only mode)
        # CREDIT parties with limit → use limit
        # CREDIT parties with no limit → None (unlimited)
        if r["payment_type"] == "CASH":
            credit_limit_val = Decimal("0")
        elif credit_limit and credit_limit > 0:
            credit_limit_val = Decimal(str(credit_limit))
        else:
            credit_limit_val = None

        # Try with original phone; fall back to placeholder if duplicate
        for attempt in range(3):
            try:
                cur.execute("""
                    INSERT INTO jc_customers
                        (business_name, person_name, phone, password_hash, secondary_phone,
                         address, city_id, credit_limit, is_active,
                         party_number, marker_1, marker_2, payment_type, notes)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    RETURNING id
                """, (
                    r["party_name"],
                    r["person"],
                    phone,
                    pw_hash,
                    r["secondary_phone"],
                    r["address"],
                    city_id,
                    credit_limit_val,
                    r["is_active"],
                    r["party_number"],
                    r["marker_1"],
                    r["marker_2"],
                    r["payment_type"],
                    r["notes"],
                ))
                customer_id = cur.fetchone()["id"]

                # Insert opening balance into AR ledger
                ob = r["opening_balance"]
                if ob != 0:
                    direction = "debit" if ob > 0 else "credit"
                    cur.execute("""
                        INSERT INTO jc_ar_ledger_entries
                            (customer_id, entry_type, amount, description, created_by_type, created_by_name)
                        VALUES (%s, 'opening_balance', %s, %s, 'admin', 'Import Script')
                    """, (
                        customer_id,
                        Decimal(str(ob)),
                        f"Opening balance [{direction} ₹{abs(ob):.2f}]",
                    ))

                imported += 1
                break
            except psycopg2.errors.UniqueViolation:
                conn.rollback()
                # Phone clash — generate a fresh placeholder
                phone = make_placeholder_phone()
                while phone in used_phones:
                    phone = make_placeholder_phone()
                used_phones.add(phone)
                pw_hash = hash_password(phone[-4:])
            except Exception as e:
                conn.rollback()
                print(f"  ERROR inserting {r['party_name']}: {e}")
                skipped += 1
                break
        else:
            print(f"  SKIP (too many phone clashes): {r['party_name']}")
            skipped += 1
            continue

        if imported % 100 == 0:
            conn.commit()
            print(f"  … {imported} inserted")

    conn.commit()
    print(f"\nDone. Imported: {imported}, Skipped: {skipped}")
    conn.close()


if __name__ == "__main__":
    main()
