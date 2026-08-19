#!/usr/bin/env python3
"""Merge ONLY DEBTOR NEW (phones) + opening balance → Cleaned / Needs Review."""

from __future__ import annotations

import re
import sys
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
PHONE_SRC = ROOT / "ONLY DEBTOR NEW .XLSX"
OB_SRC = ROOT / "debtors_opening_balance_cleaned.xlsx"
OUT = ROOT / "debtors_merged_cleaned.xlsx"

sys.path.insert(0, str(ROOT / "scripts"))
from clean_debtors import build_known_cities, extract_phones, split_name  # noqa: E402
from clean_debtors_r2 import clean_business, clean_city, merge_extra, title_case  # noqa: E402


def norm_key(name: str) -> str:
    """Normalize party name for matching across sheets."""
    s = str(name or "").strip().lower()
    s = s.replace("–", "-").replace("—", "-")
    # drop common tally tags for matching
    s = re.sub(
        r"\((cash|block|bloc|full|old|new|strict|g|j|f|h|a|w|r|jb|su|imli|\d{2}-\d{2})\)\*?",
        " ",
        s,
        flags=re.I,
    )
    s = re.sub(r"\*{1,2}", " ", s)
    s = re.sub(r"[^\w\s,&./-]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def money(v) -> str | None:
    if v is None or v == "":
        return None
    try:
        n = Decimal(str(v))
    except (InvalidOperation, ValueError):
        return None
    if n == 0:
        return "0"
    if n == n.to_integral_value():
        return format(n.quantize(Decimal("1")), "f")
    return format(n.quantize(Decimal("0.01")), "f")


def style_header(ws, widths: dict[str, int]) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def load_phones() -> dict[str, dict]:
    wb = load_workbook(PHONE_SRC, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    start = 0
    for i, r in enumerate(rows[:20]):
        vals = [str(x).lower() if x else "" for x in (r or [])]
        if "particulars" in vals or "mobile" in " ".join(vals):
            start = i + 1
            break
    out: dict[str, dict] = {}
    for r in rows[start:]:
        if not r or not r[0] or not str(r[0]).strip():
            continue
        raw = str(r[0]).strip()
        if raw.lower() in {"grand total", "total"}:
            continue
        contact = str(r[1]).strip() if r[1] and str(r[1]).strip() not in {"", "None"} else ""
        mobile = r[2]
        tel = r[3]
        email = str(r[4]).strip() if len(r) > 4 and r[4] else ""
        key = norm_key(raw)
        p, s, notes = extract_phones(mobile, tel)
        # if mobile empty but tel has mobile, extract_phones already handles both
        if not p and not s:
            p2, s2, n2 = extract_phones(tel, mobile)
            p, s, notes = p2, s2, n2
        rec = {
            "original_name": raw,
            "contact_person": contact,
            "primary_phone": p or None,
            "secondary_phone": s or None,
            "phone_notes": notes or None,
            "email": email or None,
            "mobile_raw": mobile,
            "tel_raw": tel,
        }
        # prefer record with more phone info if duplicate keys
        prev = out.get(key)
        if not prev or (p and not prev.get("primary_phone")):
            out[key] = rec
    wb.close()
    return out


def load_opening() -> dict[str, dict]:
    wb = load_workbook(OB_SRC, data_only=True)
    out: dict[str, dict] = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(h or "").lower() for h in rows[0]]
        # find cols
        def col(*names):
            for n in names:
                if n in header:
                    return header.index(n)
            return None

        i_orig = col("original_name") if col("original_name") is not None else 0
        i_biz = col("business_name")
        i_city = col("city")
        i_extra = col("extra_details")
        i_debit = col("debit")
        i_credit = col("credit")
        for r in rows[1:]:
            if not r or not r[i_orig]:
                continue
            raw = str(r[i_orig]).strip()
            key = norm_key(raw)
            rec = {
                "original_name": raw,
                "business_name": r[i_biz] if i_biz is not None else None,
                "city": r[i_city] if i_city is not None else None,
                "extra_details": r[i_extra] if i_extra is not None else None,
                "debit": money(r[i_debit]) if i_debit is not None else None,
                "credit": money(r[i_credit]) if i_credit is not None else None,
            }
            prev = out.get(key)
            if not prev:
                out[key] = rec
            else:
                # keep amounts if missing
                if not prev.get("debit") and rec.get("debit"):
                    prev["debit"] = rec["debit"]
                if not prev.get("credit") and rec.get("credit"):
                    prev["credit"] = rec["credit"]
                if len(raw) > len(prev["original_name"]):
                    prev["original_name"] = raw
    wb.close()
    return out


def review_reasons(row: dict) -> list[str]:
    rs = []
    if not row.get("business_name"):
        rs.append("no_business")
    if not row.get("city"):
        rs.append("no_city")
    if row.get("extra"):
        rs.append("has_extra")
    if not row.get("primary_phone"):
        rs.append("no_primary_phone")
    if row.get("phone_notes"):
        rs.append("phone_notes")
    if not row.get("opening_balance") and not row.get("credit"):
        rs.append("no_opening_amount")
    if row.get("match_status") == "phones_only":
        rs.append("no_opening_match")
    if row.get("match_status") == "opening_only":
        rs.append("no_phone_sheet_match")
    return rs


def main() -> None:
    phones = load_phones()
    opening = load_opening()

    phone_keys = set(phones)
    ob_keys = set(opening)
    both = phone_keys & ob_keys
    only_phone = phone_keys - ob_keys
    only_ob = ob_keys - phone_keys

    print(f"Phone sheet parties: {len(phones)}")
    print(f"Opening sheet parties: {len(opening)}")
    print(f"Matched (same normalized name): {len(both)}")
    print(f"Phones only (no opening): {len(only_phone)}")
    print(f"Opening only (no phones sheet): {len(only_ob)}")

    # build known cities from all original names
    all_names = [phones[k]["original_name"] for k in phones] + [opening[k]["original_name"] for k in opening]
    known = build_known_cities(all_names)

    merged: list[dict] = []
    all_keys = sorted(phone_keys | ob_keys, key=lambda k: (phones.get(k) or opening.get(k) or {}).get("original_name", k).lower())

    for key in all_keys:
        p = phones.get(key)
        o = opening.get(key)
        if p and o:
            match_status = "matched"
            raw = o["original_name"] if len(o["original_name"]) >= len(p["original_name"]) else p["original_name"]
            # prefer OB original if it has more tags, else phone
            if "(" in o["original_name"] and "(" not in p["original_name"]:
                raw = o["original_name"]
            elif len(p["original_name"]) > len(o["original_name"]):
                raw = p["original_name"]
            else:
                raw = o["original_name"]
        elif p:
            match_status = "phones_only"
            raw = p["original_name"]
        else:
            match_status = "opening_only"
            raw = o["original_name"]

        biz, city, extra = split_name(raw, known)
        city2, city_extra = clean_city(city, known)
        biz2, maybe_city, biz_extra = clean_business(biz, city2, known)
        if maybe_city and not city2:
            city2 = maybe_city

        # if OB already had cleaned fields and city empty, try those
        if o and not city2 and o.get("city"):
            city2, more = clean_city(str(o["city"]), known)
            city_extra = merge_extra(city_extra, more)
        if o and o.get("extra_details"):
            extra = merge_extra(extra, o["extra_details"])
        if o and o.get("business_name") and (not biz2 or len(str(o["business_name"])) > len(biz2)):
            # keep our cleaned biz unless OB biz looks better and ours empty
            if not biz2:
                biz2 = title_case(str(o["business_name"]))

        extras = [extra, city_extra, biz_extra]
        if p and p.get("contact_person"):
            extras.append(f"contact:{p['contact_person']}")
        if p and p.get("email"):
            extras.append(f"email:{p['email']}")

        primary = p.get("primary_phone") if p else None
        secondary = p.get("secondary_phone") if p else None
        phone_notes = p.get("phone_notes") if p else None

        # opening: debit = opening balance due; credit = credit balance
        debit = o.get("debit") if o else None
        credit = o.get("credit") if o else None

        row = {
            "original_name": raw,
            "business_name": biz2 or None,
            "city": city2 or None,
            "extra": merge_extra(*[x for x in extras if x]) or None,
            "primary_phone": primary,
            "secondary_phone": secondary,
            "opening_balance": debit,  # debit due
            "credit": credit,
            "phone_notes": phone_notes,
            "match_status": match_status,
        }
        merged.append(row)

    cols = [
        "original_name",
        "business_name",
        "city",
        "extra",
        "primary_phone",
        "secondary_phone",
        "opening_balance",
        "credit",
    ]

    out = Workbook()
    # match report sheet first? user asked 2 sheets only - put report as 3rd small or print only
    cleaned = out.active
    cleaned.title = "Cleaned"
    cleaned.append(cols)
    review = out.create_sheet("Needs Review")
    review.append(cols + ["review_reason", "match_status"])
    match_sheet = out.create_sheet("Match Report")
    match_sheet.append(["metric", "count"])
    match_sheet.append(["phone_sheet_parties", len(phones)])
    match_sheet.append(["opening_sheet_parties", len(opening)])
    match_sheet.append(["matched", len(both)])
    match_sheet.append(["phones_only", len(only_phone)])
    match_sheet.append(["opening_only", len(only_ob)])
    match_sheet.append(["", ""])
    match_sheet.append(["phones_only_sample", ""])
    for k in sorted(only_phone, key=lambda x: phones[x]["original_name"].lower())[:40]:
        match_sheet.append([phones[k]["original_name"], "phones_only"])
    match_sheet.append(["", ""])
    match_sheet.append(["opening_only_sample", ""])
    for k in sorted(only_ob, key=lambda x: opening[x]["original_name"].lower())[:40]:
        match_sheet.append([opening[k]["original_name"], "opening_only"])

    n_c = n_r = 0
    for row in merged:
        rs = review_reasons(row)
        vals = [row[c] for c in cols]
        if rs:
            review.append(vals + [", ".join(rs), row["match_status"]])
            n_r += 1
        else:
            cleaned.append(vals)
            n_c += 1

    style_header(cleaned, {
        "A": 44, "B": 28, "C": 20, "D": 28, "E": 14, "F": 14, "G": 16, "H": 12,
    })
    style_header(review, {
        "A": 44, "B": 28, "C": 20, "D": 28, "E": 14, "F": 14, "G": 16, "H": 12, "I": 32, "J": 14,
    })

    out.save(OUT)
    print(f"\nWrote {OUT}")
    print(f"Cleaned (sure): {n_c}")
    print(f"Needs Review: {n_r}")
    print(f"Total: {n_c + n_r}")


if __name__ == "__main__":
    main()
