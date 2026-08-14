"""Excel / CSV export builders + full backup zip."""

from __future__ import annotations

import csv
import io
import zipfile
from datetime import datetime, timezone
from decimal import Decimal
from typing import Any, Iterable, List, Sequence

from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.models.activity_log import ActivityLog
from app.models.catalog_product import CatalogProduct
from app.models.customer import Customer
from app.models.customer_bill import CustomerBill
from app.models.expense import Expense
from app.models.stock import StockBalance
from app.models.vendor import Vendor
from app.services.ap_ledger import list_ap_vendors
from app.services.ar_ledger import list_ar_customers


def _money(v: Any) -> str:
    if v is None:
        return ""
    try:
        return format(Decimal(str(v)), "f")
    except Exception:
        return str(v)


def workbook_from_rows(sheet_name: str, headers: Sequence[str], rows: Iterable[Sequence[Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Sheet1")[:31]
    ws.append(list(headers))
    for row in rows:
        ws.append(["" if c is None else c for c in row])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def multi_sheet_xlsx(sheets: List[tuple[str, Sequence[str], List[Sequence[Any]]]]) -> bytes:
    wb = Workbook()
    first = True
    for name, headers, rows in sheets:
        if first:
            ws = wb.active
            first = False
        else:
            ws = wb.create_sheet()
        ws.title = (name or "Sheet")[:31]
        ws.append(list(headers))
        for row in rows:
            ws.append(["" if c is None else c for c in row])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_customers(db: Session) -> bytes:
    rows = (
        db.query(Customer)
        .filter(Customer.deleted_at.is_(None))
        .order_by(Customer.business_name)
        .all()
    )
    return workbook_from_rows(
        "Customers",
        ["id", "business_name", "person_name", "phone", "gst", "credit_limit", "credit_override", "city_id", "route_id", "address"],
        [
            [
                c.id, c.business_name, c.person_name, c.phone, c.gst_number,
                _money(c.credit_limit), bool(c.credit_override), c.city_id, c.route_id, c.address,
            ]
            for c in rows
        ],
    )


def export_vendors(db: Session) -> bytes:
    rows = db.query(Vendor).filter(Vendor.deleted_at.is_(None)).order_by(Vendor.business_name).all()
    return workbook_from_rows(
        "Vendors",
        ["id", "business_name", "person_name", "phone", "gst", "address"],
        [[v.id, v.business_name, v.person_name, v.phone, v.gst_number, v.address] for v in rows],
    )


def export_catalog(db: Session) -> bytes:
    rows = (
        db.query(CatalogProduct)
        .filter(CatalogProduct.deleted_at.is_(None))
        .order_by(CatalogProduct.our_product_id)
        .limit(20000)
        .all()
    )
    return workbook_from_rows(
        "Catalog",
        ["id", "our_product_id", "vendor_id", "buying_price", "selling_price", "unit", "category"],
        [
            [
                p.id, p.our_product_id, p.vendor_id,
                _money(p.buying_price),
                _money(p.selling_price),
                p.unit or "",
                p.category or "",
            ]
            for p in rows
        ],
    )


def export_stock(db: Session) -> bytes:
    bals = db.query(StockBalance).all()
    ids = [b.catalog_product_id for b in bals]
    prod = {}
    if ids:
        prod = {p.id: p for p in db.query(CatalogProduct).filter(CatalogProduct.id.in_(ids)).all()}
    return workbook_from_rows(
        "Stock",
        ["catalog_product_id", "our_product_id", "qty_on_hand", "threshold", "selling_price"],
        [
            [
                b.catalog_product_id,
                (prod[b.catalog_product_id].our_product_id if b.catalog_product_id in prod else ""),
                b.quantity_on_hand,
                b.low_stock_threshold,
                _money(prod[b.catalog_product_id].selling_price) if b.catalog_product_id in prod else "",
            ]
            for b in bals
        ],
    )


def export_ar(db: Session) -> bytes:
    items = list_ar_customers(db)
    return workbook_from_rows(
        "AR",
        ["customer_id", "label", "outstanding", "opening", "bills", "payments", "credits", "txns"],
        [
            [
                r["customer_id"], r["customer_label"], r["outstanding"], r.get("opening_total"),
                r["bill_total"], r["payment_total"], r.get("credit_total"), r.get("transaction_count"),
            ]
            for r in items
        ],
    )


def export_ap(db: Session) -> bytes:
    items = list_ap_vendors(db)
    return workbook_from_rows(
        "AP",
        ["vendor_id", "label", "outstanding", "opening", "bills", "payments", "debits", "txns"],
        [
            [
                r["vendor_id"], r["vendor_label"], r["outstanding"], r.get("opening_total"),
                r["bill_total"], r["payment_total"], r.get("debit_note_total"), r.get("transaction_count"),
            ]
            for r in items
        ],
    )


def export_expenses(db: Session) -> bytes:
    rows = db.query(Expense).order_by(Expense.expense_date.desc()).limit(10000).all()
    return workbook_from_rows(
        "Expenses",
        ["id", "date", "category", "amount", "description", "reference", "by"],
        [
            [e.id, e.expense_date.isoformat(), e.category, _money(e.amount), e.description, e.reference, e.created_by_name]
            for e in rows
        ],
    )


def export_sales_bills(db: Session) -> bytes:
    rows = db.query(CustomerBill).order_by(CustomerBill.id.desc()).limit(20000).all()
    return workbook_from_rows(
        "SalesBills",
        ["id", "bill_number", "customer_id", "created_at", "grand_total", "gst_amount"],
        [
            [
                b.id, b.bill_number, b.customer_id,
                b.created_at.isoformat() if b.created_at else "",
                _money(b.grand_total),
                _money(b.gst_amount),
            ]
            for b in rows
        ],
    )


def export_activity(db: Session) -> bytes:
    rows = db.query(ActivityLog).order_by(ActivityLog.id.desc()).limit(20000).all()
    return workbook_from_rows(
        "Activity",
        ["id", "at", "actor", "action", "entity_type", "entity_id", "label", "detail"],
        [
            [
                a.id,
                a.created_at.isoformat() if a.created_at else "",
                a.actor_name,
                a.action,
                a.entity_type,
                a.entity_id,
                a.entity_label,
                a.detail,
            ]
            for a in rows
        ],
    )


EXPORT_KINDS = {
    "customers": export_customers,
    "vendors": export_vendors,
    "catalog": export_catalog,
    "stock": export_stock,
    "ar": export_ar,
    "ap": export_ap,
    "expenses": export_expenses,
    "sales_bills": export_sales_bills,
    "activity": export_activity,
}


def build_backup_zip(db: Session) -> bytes:
    buf = io.BytesIO()
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for kind, fn in EXPORT_KINDS.items():
            zf.writestr(f"{kind}.xlsx", fn(db))
        zf.writestr(
            "manifest.txt",
            f"JC backup {stamp} UTC\nSheets: {', '.join(EXPORT_KINDS)}\n",
        )
    return buf.getvalue()


def csv_bytes(headers: Sequence[str], rows: Iterable[Sequence[Any]]) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(list(headers))
    for row in rows:
        w.writerow(row)
    return buf.getvalue().encode("utf-8-sig")
