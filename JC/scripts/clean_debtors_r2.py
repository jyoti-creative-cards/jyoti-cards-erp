#!/usr/bin/env python3
"""Round-2 deep clean of debtors_cleaned.xlsx (after manual edits)."""

from __future__ import annotations

import re
import shutil
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "debtors_cleaned.xlsx"
BAK = ROOT / "debtors_cleaned_before_r2.xlsx"
OUT = ROOT / "debtors_cleaned.xlsx"

NOISE_IN_CITY = re.compile(
    r"""(?ix)
    \b(
        bus\s*only|parcel\s*only|only|direct|trans|not\s*billing|billing|
        new\s*due|full\s*bill|strict|strct|strtct|stct|block|bloc|cash|
        imli|3imli|teenimli|sample\s*print
    )\b
    |
    \d+\s*/- |
    \d{2}-\d{2} |
    \*{1,2}
    """
)

FIRST_NAMES = {
    "sanjay", "ramesh", "mukesh", "rajesh", "suresh", "anil", "sunil", "amit",
    "rahul", "vikas", "deepak", "prakash", "mahesh", "dinesh", "naresh", "paresh",
    "hitesh", "jitin", "jeevan", "kishor", "dilip", "santosh", "vishnu", "gurjar",
    "savita", "swati", "aman", "imran", "mustafa", "husen", "ronak", "seenu",
    "bunty", "chakor", "ankur", "anshul", "kamlesh", "kailash", "paramjit",
    "yaqoob", "riyaz", "gopu", "shekhar", "vishwas", "pandya", "pancholi",
    "harpal", "pankaj", "bhagel", "baghel", "govind", "uday", "ishwar",
    "santosh", "mohammad", "manni", "mama", "himanshu",
}

NOT_A_CITY = {
    "shop", "sales", "self", "book", "store", "tower", "road", "square",
    "freind", "friend", "aaron", "19", "su", "g", "j", "r",
    "h", "f", "a", "w", "jb", "imli",
}

CITY_ALIASES = {
    "badwani": "Barwani",
    "barwani": "Barwani",
    "amerli": "Amreli",
    "amreli": "Amreli",
    "ashok nagar": "Ashoknagar",
    "ashoknagar": "Ashoknagar",
    "akrela": "Akrela",
    "aklera": "Akrela",
    "narsingarh": "Narsinghgarh",
    "narsullaganj": "Narsulaganj",
    "narsulaganj": "Narsulaganj",
    "dhamnaud": "Dhamnod",
    "dhamnod": "Dhamnod",
    "seoni malwa": "Seoni Malwa",
    "jhalrapatan": "Jhalrapatan",
    "jhalawad": "Jhalawar",
    "jhalawar": "Jhalawar",
    "baswada": "Banswara",
    "banswada": "Banswara",
    "banswara": "Banswara",
    "gujrat": "Gujarat",
    "gujarat": "Gujarat",
    "rajasthan": "Rajasthan",
    "rj": "Rajasthan",
    "raj.": "Rajasthan",
    "mah.": "Maharashtra",
    "maharashtra": "Maharashtra",
    "baroda": "Vadodara",
    "mow": "Mhow",
    "mhow": "Mhow",
    "indoree": "Indore",
    "aalot": "Alot",
    "alot": "Alot",
    "biora": "Biaora",
    "biaora": "Biaora",
    "pidawa": "Pidawa",
    "soyatkala": "Soyat Kala",
    "soyat kala": "Soyat Kala",
    "khajrana": "Khajrana",
    "gautampura": "Gautampura",
    "deapalpur": "Depalpur",
    "depalpur": "Depalpur",
    "nogaon": "Naugaon",
    "palyamataimli": "Palyamata",
    "palyamata": "Palyamata",
    "pratap garh": "Pratapgarh",
    "pratapgarh": "Pratapgarh",
    "arond": "Arond",
    "udaynagar": "Udaynagar",
}


def title_token(tok: str) -> str:
    if not tok:
        return tok
    if re.fullmatch(r"(?:[A-Za-z]\.){2,}", tok) or re.fullmatch(r"[A-Za-z](?:\.[A-Za-z])+\.?", tok):
        return tok.upper()
    if tok.isupper() and len(tok) <= 3 and tok.isalpha():
        return tok.upper()
    if "-" in tok and not tok.startswith("-"):
        return "-".join(title_token(p) if p else p for p in tok.split("-"))
    return tok[:1].upper() + tok[1:].lower() if len(tok) > 1 else tok.upper()


def title_case(name: str) -> str:
    name = re.sub(r"\s+", " ", (name or "").strip(" -–,\t"))
    if not name:
        return ""
    parts = [title_token(t) for t in name.split(" ") if t]
    out = " ".join(parts)
    out = re.sub(r"\s+&", " &", out)
    out = re.sub(r"\s+,", ",", out)
    return out


def alias_city(name: str) -> str:
    key = re.sub(r"\s+", " ", name.strip()).lower()
    if key in CITY_ALIASES:
        return CITY_ALIASES[key]
    return title_case(name)


def merge_extra(*parts: str) -> str:
    seen = set()
    out = []
    for p in parts:
        if not p:
            continue
        for bit in re.split(r"\s*\|\s*", str(p)):
            bit = re.sub(r"\s+", " ", bit).strip(" -–,")
            if not bit:
                continue
            k = bit.lower()
            if k in seen:
                continue
            seen.add(k)
            out.append(bit)
    return " | ".join(out)


def extract_parens(text: str) -> tuple[str, list[str]]:
    tags = []
    text = text or ""

    def _g(m: re.Match) -> str:
        tags.append(m.group(1).strip())
        return " "

    cleaned = re.sub(r"\(([^)]*)\)", _g, text)
    # unclosed trailing paren: "Meghnagar (bloc By Jeevan B"
    m = re.search(r"\(([^)]*)$", cleaned)
    if m:
        tags.append(m.group(1).strip())
        cleaned = cleaned[: m.start()]
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" -–,/")
    return cleaned, tags


def build_known_cities(cities: list[str]) -> set[str]:
    c = Counter()
    for city in cities:
        if not city:
            continue
        base, _ = extract_parens(str(city))
        base = re.sub(r"[*]+", "", base)
        base = re.sub(r"\d+", " ", base)
        base = re.sub(r"\s+", " ", base).strip().lower()
        if base:
            c[base] += 1
            # also first token
            c[base.split()[0]] += 1
    known = {k for k, v in c.items() if v >= 2 and len(k) >= 3}
    known |= set(CITY_ALIASES.keys())
    known |= {
        "indore", "dewas", "ratlam", "ujjain", "dhar", "manawar", "khargone",
        "khandwa", "bhopal", "jhabua", "alirajpur", "pithampur", "mhow",
        "raipur", "rajkot", "pune", "surat", "nagpur", "jodhpur", "kota",
        "betul", "sehore", "jabalpur", "barwani", "sanawad", "thandla",
        "petlawad", "kukshi", "maheshwar", "ashta", "jaora", "rau", "dhule",
        "beawar", "mumbai", "sendhwa", "gandhwani", "meghnagar", "barnagar",
        "depalpur", "gautampura", "burhanpur", "itarsi", "narsinghgarh",
        "palyamata", "nisarpur", "sukhliya", "biaora", "alot", "dhamnod",
    }
    return known


def looks_like_real_city(token: str, known: set[str]) -> bool:
    t = re.sub(r"\s+", " ", (token or "").strip())
    if not t:
        return False
    key = t.lower()
    if key in NOT_A_CITY or key in FIRST_NAMES:
        return False
    if re.search(r"[\u0900-\u097F]", t):
        return False
    if re.search(r"\d|\bdue\b|\bparcel\b|\bonly\b|\bdirect\b|\btrans\b|\btower\b|\bprinters?\b|\bshop\b|\bfriends?\b|\bfreind\b|\brs\.?\b", t, re.I):
        return False
    if key.startswith("/") or key.startswith("-"):
        return False
    if len(key) <= 2 and key not in {"rj", "mp"}:
        return False
    parts = key.split()
    if any(p in FIRST_NAMES for p in parts):
        return False
    if any(p in {"road", "tower", "square", "chowki", "phatak", "bus", "only"} for p in parts):
        # "Petlawad Road" alone is not a city — parent place handled elsewhere
        if not (len(parts) >= 2 and parts[0] in known):
            return False
    if key in known or parts[0] in known:
        return True
    # multi-word place like Bhawani Mandi / Seoni Malwa
    if len(parts) == 2 and all(len(p) >= 3 for p in parts) and parts[-1] not in {"ji", "tr", "co"}:
        return True
    if len(parts) == 1 and len(parts[0]) >= 4 and parts[0].isalpha():
        return True
    return False


def clean_city(city: str, known: set[str]) -> tuple[str, str]:
    """Return (city, extra_bits)."""
    if not city:
        return "", ""
    raw = str(city).strip()
    extras: list[str] = []

    # strip leading slash / normalize slash city pairs
    raw = re.sub(r"^[/\\]+\s*", "", raw)
    # "Ratlam / Not" / "Udaynagar/ Indore"
    if "/" in raw:
        left, right = [x.strip() for x in raw.split("/", 1)]
        if re.search(r"^(not|no|nil)\b", right, re.I):
            extras.append(right)
            raw = left
        elif looks_like_real_city(left, known) and looks_like_real_city(right, known):
            raw = f"{left} ({right})"
        elif looks_like_real_city(left, known):
            if right:
                extras.append(right)
            raw = left
        elif looks_like_real_city(right, known):
            if left:
                extras.append(left)
            raw = right
    # Devanagari chunks → extra
    dev = re.findall(r"[\u0900-\u097F]+", raw)
    for d in dev:
        extras.append(d)
    raw = re.sub(r"[\u0900-\u097F]+", " ", raw)

    base, tags = extract_parens(raw)

    # glue fixes
    base_l = base.lower().replace(" ", "")
    if base_l.endswith("imli") and len(base_l) > 4:
        # Palyamataimli / Teenimli / 3imli already in base
        m = re.match(r"^(.*?)(\d*imli)$", base, re.I)
        if m and m.group(1):
            extras.append(m.group(2))
            base = m.group(1).strip()
    if re.search(r"\bteenimli\b", base, re.I):
        extras.append("Teenimli")
        base = re.sub(r"\bteenimli\b", " ", base, flags=re.I)
    if re.search(r"\b3imli\b", base, re.I):
        extras.append("3Imli")
        base = re.sub(r"\b3imli\b", " ", base, flags=re.I)

    # pull noise words from base into extra
    for m in list(NOISE_IN_CITY.finditer(base)):
        extras.append(m.group(0).strip())
    base = NOISE_IN_CITY.sub(" ", base)
    base = re.sub(r"[*]+", " ", base)
    base = re.sub(r"['\"]+", " ", base)
    base = re.sub(r"\s+", " ", base).strip(" -–,/")

    # trailing bare numbers (Manpur 5, Kesur 10, Burhanpur 15)
    mnum = re.match(r"^(.*?)[\s-]+(\d{1,4})$", base)
    if mnum and looks_like_real_city(mnum.group(1), known):
        extras.append(mnum.group(2))
        base = mnum.group(1).strip()

    # MR10 / area codes → prefer paren city
    if re.fullmatch(r"mr\s*10", base, re.I) or base.lower() in {"mr10"}:
        extras.append("MR10")
        base = ""

    # "Petlawad Road (Thandla)" style already split; bare "X Road"/"Tower"
    if re.search(r"\b(road|tower|square|chowki|phatak)\b", base, re.I):
        # keep leading place token if known: "Petlawad Road" → try Petlawad
        parts = base.split()
        if parts and looks_like_real_city(parts[0], known):
            extras.append(base)
            base = parts[0]
        else:
            extras.append(base)
            base = ""

    # business text wrongly in city — try "Name - City" recovery
    if re.search(r"\b(printers?|press|graphics|cards?|photocopy|studio)\b", base, re.I):
        m = re.search(r"^(.*?)\s*-\s*(.+)$", base)
        if m and looks_like_real_city(m.group(2), known):
            extras.append(m.group(1).strip())
            base = m.group(2).strip()
        else:
            extras.append(base)
            base = ""

    # person as city
    if base.lower() in FIRST_NAMES or base.lower() in NOT_A_CITY:
        extras.append(base)
        base = ""

    place_tags = []
    for tag in tags:
        t = re.sub(r"[*]+", "", tag).strip()
        if not t:
            continue
        tl = t.lower()
        if tl in {"1st", "first", "only", "new", "old", "full", "block", "cash", "su", "ak", "g", "j", "jb"}:
            extras.append(t)
            continue
        if NOISE_IN_CITY.search(t) or re.search(r"\d+\s*/-", t):
            extras.append(t)
            continue
        if any(w in FIRST_NAMES for w in tl.split()) or re.search(r"\b(tr\.?|ji|bhai|direct|harpal|pankaj|jeevan)\b", tl):
            extras.append(t)
            continue
        tl_norm = re.sub(r"\s+", " ", tl)
        if (
            looks_like_real_city(t, known)
            or tl_norm in known
            or tl_norm in CITY_ALIASES
            or tl_norm.replace(" ", "") in {k.replace(" ", "") for k in CITY_ALIASES}
        ):
            place_tags.append(alias_city(t))
        elif len(t.split()) <= 3 and t.replace(" ", "").isalpha() and len(t) >= 4:
            # locality / district style qualifier
            place_tags.append(alias_city(t))
        else:
            extras.append(t)

    city = alias_city(base) if base else ""
    # if city empty but place tag exists, promote first place tag
    if not city and place_tags:
        city = place_tags.pop(0)
    elif city and place_tags:
        # keep locality qualifiers that are real places
        for pt in place_tags:
            if pt.lower() == city.lower():
                continue
            city = f"{city} ({pt})"
    elif not city and not place_tags:
        city = ""

    # final sanity — try first token rescue before dropping
    if city:
        core = re.sub(r"\([^)]*\)", "", city).strip()
        if not looks_like_real_city(core, known) and core.lower() not in known and core.lower() not in CITY_ALIASES:
            tokens = re.split(r"[\s,/|-]+", core)
            rescued = ""
            for tok in tokens:
                if looks_like_real_city(tok, known):
                    rescued = alias_city(tok)
                    break
            if rescued:
                leftover = core.replace(tokens[0] if False else "", "").strip()
                # keep non-rescued bits as extra
                rest = " ".join(t for t in tokens if alias_city(t).lower() != rescued.lower())
                if rest:
                    extras.append(rest)
                city = rescued
            else:
                extras.insert(0, city)
                city = ""

    # cleanup city formatting
    if city:
        core, pts = extract_parens(city)
        core = alias_city(core)
        pts = [alias_city(p) for p in pts]
        city = core
        for p in pts:
            if p and p.lower() != core.lower():
                city = f"{city} ({p})"

    return city, merge_extra(*extras)


def clean_business(biz: str, city: str, known: set[str]) -> tuple[str, str, str]:
    """Return business, maybe_city_from_biz, extra."""
    extras: list[str] = []
    maybe_city = ""
    b = str(biz or "").strip()
    if not b:
        return "", "", ""

    # pull parenthetical notes from business
    base, tags = extract_parens(b)
    for tag in tags:
        tl = tag.lower().strip()
        if looks_like_real_city(tag, known) and not city:
            maybe_city = alias_city(tag)
        elif any(w in FIRST_NAMES for w in tl.split()) or re.search(r"\b(ji|bhai|trading|shri)\b", tl):
            extras.append(tag)
        else:
            extras.append(tag)
    b = base

    # trailing " - Road" leftovers
    b = re.sub(r"[\s-]+thandla\s+road\s*$", "", b, flags=re.I)
    if re.search(r"\broad\s*$", b, re.I) and "printing" not in b.lower():
        m = re.search(r"^(.*?)[\s,|-]+(\S+\s+road)\s*$", b, re.I)
        if m:
            extras.append(m.group(2).strip())
            b = m.group(1).strip()

    # "Rahul Jaishwal Manawar" — last token city if matches current city or known
    words = b.split()
    if len(words) >= 3:
        last = words[-1]
        if looks_like_real_city(last, known) and (not city or last.lower() == re.sub(r"\([^)]*\)", "", city).strip().lower() or last.lower() in known):
            # only strip if last is clearly a city and rest looks like person name
            rest = " ".join(words[:-1])
            if any(w.lower() in FIRST_NAMES for w in words[:-1]) or re.search(r"\b(ji|bhai|patel|mehta)\b", rest, re.I):
                if not city:
                    maybe_city = alias_city(last)
                elif last.lower() == re.sub(r"\([^)]*\)", "", city).strip().lower():
                    b = rest
                else:
                    extras.append(last)
                    b = rest

    # "Jeevan Jeneral Store Indore" with city wrongly "Store Indore"
    if words and words[-1].lower() in known and len(words) >= 3:
        if words[-1].lower() != "store":
            # if city empty or city looks wrong
            pass

    # normalize showroom / photocopy spelling leftovers
    b = re.sub(r"(?i)\bshow-?room\b", "Showroom", b)
    b = re.sub(r"(?i)\bmoblie\b", "Mobile", b)
    b = re.sub(r"(?i)\bjeneral\b", "General", b)
    b = re.sub(r"(?i)\bfreind\b", "Friend", b)
    b = re.sub(r"\s+", " ", b).strip(" -–,")

    # person-only ledger names: keep, but title-case; move dangling codes
    b = title_case(b)
    # fix initials spacing A.B → keep
    return b, maybe_city, merge_extra(*extras)


def main() -> None:
    shutil.copy2(SRC, BAK)
    wb = load_workbook(SRC, data_only=True)
    if "Cleaned" in wb.sheetnames:
        ws = wb["Cleaned"]
    elif "Debtors Cleaned" in wb.sheetnames:
        ws = wb["Debtors Cleaned"]
    else:
        ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h) if h else "" for h in rows[0]]
    data = rows[1:]
    # skip title-ish first rows if present
    if header and "business" not in header[0].lower() and "original" not in header[0].lower():
        # find header row
        for i, r in enumerate(rows[:5]):
            vals = [str(x or "").lower() for x in r]
            if any("business" in v for v in vals) or any("original" in v for v in vals):
                header = [str(h) if h else "" for h in rows[i]]
                data = rows[i + 1 :]
                break

    # column map
    def col(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return None

    i_orig = col("original_name") or 0
    i_biz = col("business_name") or 1
    i_city = col("city") or 2
    i_extra = col("extra_details") or 3
    i_p = col("primary_phone") or 4
    i_s = col("secondary_phone") or 5
    i_pn = col("phone_notes") or 6

    known = build_known_cities([r[i_city] for r in data if r[i_city]])

    # keep original tally sheet if present
    orig_path = ROOT / "debtors .xlsx"
    out_wb = Workbook()
    orig_ws = out_wb.active
    orig_ws.title = "Original"
    if orig_path.exists():
        owb = load_workbook(orig_path, data_only=True)
        for r in owb.active.iter_rows(values_only=True):
            orig_ws.append(list(r))
    else:
        orig_ws.append(["original_name"])
        for r in data:
            orig_ws.append([r[i_orig]])

    out = out_wb.create_sheet("Cleaned")
    out_cols = [
        "original_name", "business_name", "city", "extra_details",
        "primary_phone", "secondary_phone", "phone_notes",
    ]
    out.append(out_cols)
    rev = out_wb.create_sheet("Needs Review")
    rev.append(out_cols + ["review_reason"])

    moved_city = 0
    cleaned_rows: list[list] = []
    for r in data:
        orig = r[i_orig]
        biz = r[i_biz]
        city = r[i_city]
        extra = r[i_extra]
        pp, sp, pn = r[i_p], r[i_s], r[i_pn]
        notes = []

        city2, city_extra = clean_city(str(city or ""), known)
        if (city or "") and city2 != str(city or "").strip():
            notes.append("city_cleaned")
        if city_extra and not city2 and (city or ""):
            moved_city += 1
            notes.append("city_to_extra")

        # promote place-like bits already sitting in extra_details into city qualifier
        kept_extra = []
        skip_prom = {
            "billing", "block", "bloc", "cash", "retail", "new", "old", "full",
            "strict", "only", "not", "imli", "su", "jb", "direct",
        }
        for bit in re.split(r"\s*\|\s*", str(extra or "")):
            bit = bit.strip()
            if not bit:
                continue
            bl = re.sub(r"\s+", " ", bit.lower())
            if bl in skip_prom or re.search(r"strict|block|cash|billing|due|parcel|\bonly\b", bl):
                kept_extra.append(bit)
                continue
            if city2 and (
                bl in CITY_ALIASES
                or bl in known
                or bl.replace(" ", "") in {k.replace(" ", "") for k in CITY_ALIASES}
                or (len(bit.split()) <= 2 and bit.replace(" ", "").isalpha() and looks_like_real_city(bit, known))
            ):
                qual = alias_city(bit)
                core = re.sub(r"\([^)]*\)", "", city2).strip().lower()
                if qual.lower() != core and f"({qual.lower()})" not in city2.lower():
                    city2 = f"{city2} ({qual})"
                    notes.append("extra_to_city_qual")
                continue
            if not city2 and looks_like_real_city(bit, known):
                city2 = alias_city(bit)
                notes.append("city_from_extra")
                continue
            kept_extra.append(bit)
        extra = " | ".join(kept_extra)

        biz2, maybe_city, biz_extra = clean_business(str(biz or ""), city2, known)
        if maybe_city and not city2:
            city2 = maybe_city
            notes.append("city_from_biz")
        if biz2 != title_case(str(biz or "")) and str(biz or "").strip():
            # still flag meaningful biz change
            if biz2.lower() != str(biz or "").strip().lower():
                notes.append("biz_cleaned")

        # empty business but city looks like business name
        if not biz2 and city2 and re.search(r"\b(printers?|press|graphics|cards?)\b", city2, re.I):
            biz2 = title_case(re.sub(r"\([^)]*\)", "", city2).split("-")[0].strip())
            # try extract city from original
            city2, more = clean_city(str(city or ""), known)
            extra = merge_extra(str(extra or ""), city_extra, more)
            notes.append("biz_from_city")

        # Mama + Bhanja Printers in extra
        if str(biz2).lower() in {"mama"} and re.search(r"bhanja\s*printers", str(city_extra or ""), re.I):
            biz2 = "Bhanja Printers"
            city_extra = re.sub(r"(?i)bhanja\s*printers\s*-?\s*", "", city_extra or "").strip(" |")
            notes.append("split_mama_bhanja")

        # "Jaipur" as business + "Bombay Cards" in extra → swap
        if re.search(r"\b(cards?|press|printers?|graphics|photocopy)\b", str(city_extra or ""), re.I):
            if looks_like_real_city(str(biz2 or ""), known) and not city2:
                city2 = alias_city(biz2)
                m = re.search(r"([A-Za-z].*(?:cards?|press|printers?|graphics|photocopy)[A-Za-z\s.]*)", str(city_extra), re.I)
                if m:
                    biz2 = title_case(m.group(1))
                    city_extra = merge_extra(re.sub(re.escape(m.group(1)), "", str(city_extra), flags=re.I))
                    notes.append("swap_city_biz")

        # Deepak Shop / Jeevan Friend sitting only in extra
        if not biz2 and city_extra and not city2:
            biz2 = title_case(re.split(r"\s*\|\s*", city_extra)[0])
            city_extra = merge_extra(*re.split(r"\s*\|\s*", city_extra)[1:])
            notes.append("biz_from_extra")

        extra2 = merge_extra(str(extra or ""), city_extra, biz_extra)

        # phone keep as-is (already cleaned r1); normalize empty
        pp = str(pp).strip() if pp not in (None, "") else ""
        sp = str(sp).strip() if sp not in (None, "") else ""
        pn = str(pn).strip() if pn not in (None, "") else ""
        if pp.lower() == "none":
            pp = ""
        if sp.lower() == "none":
            sp = ""

        row_out = [orig, biz2, city2, extra2 or None, pp or None, sp or None, pn or None]
        out.append(row_out)
        cleaned_rows.append(row_out)

        reasons = []
        if not biz2:
            reasons.append("no_business")
        if not city2:
            reasons.append("no_city")
        if extra2:
            reasons.append("has_extra")
        if not pp:
            reasons.append("no_primary_phone")
        if pn:
            reasons.append("phone_notes")
        if "city_to_extra" in notes:
            reasons.append("city_moved")
        if reasons:
            rev.append(row_out + [", ".join(reasons)])

    out_wb.save(OUT)
    print(f"Backup: {BAK}")
    print(f"Wrote: {OUT}")
    print(f"Sheets: Original / Cleaned / Needs Review")
    print(f"Rows: {len(cleaned_rows)}")
    print(f"Cities moved to extra: {moved_city}")
    print(f"Review rows: {rev.max_row - 1}")


if __name__ == "__main__":
    main()
