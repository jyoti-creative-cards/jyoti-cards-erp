#!/usr/bin/env python3
"""Compare ONLY DEBTOR NEW sheet to jc_customers. Print missing, cleaned like the app."""
from __future__ import annotations

import json
import os
import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "ONLY DEBTOR NEW.XLSX"
OUT_XLSX = ROOT / "debtors_missing_from_db.xlsx"
OUT_JSON = ROOT / "debtors_missing_from_db.json"
BACKEND = ROOT / "backend"

sys.path.insert(0, str(ROOT / "scripts"))
sys.path.insert(0, str(BACKEND))
os.chdir(BACKEND)

from clean_debtors import build_known_cities, extract_phones, split_name  # noqa: E402
from clean_debtors_r2 import clean_business, clean_city, merge_extra, title_case  # noqa: E402
from merge_debtors_final import norm_key  # noqa: E402

from app.db.session import SessionLocal, init_db  # noqa: E402
from app.models.city import City  # noqa: E402
from app.models.customer import Customer  # noqa: E402
from app.models.route import Route  # noqa: E402


def digits_phone(v) -> str | None:
    if v is None:
        return None
    d = re.sub(r"\D+", "", str(v))
    if len(d) == 11 and d.startswith("0"):
        d = d[1:]
    if len(d) == 12 and d.startswith("91"):
        d = d[2:]
    if len(d) == 10 and d[0] in "6789":
        return d
    return None


def load_sheet() -> list[dict]:
    wb = load_workbook(SRC, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    start = 0
    for i, r in enumerate(rows[:20]):
        vals = [str(x).lower() if x else "" for x in (r or [])]
        if "particulars" in vals:
            start = i + 1
            break
    names = []
    raw_rows = []
    for r in rows[start:]:
        if not r or not r[0] or not str(r[0]).strip():
            continue
        raw = str(r[0]).strip()
        if raw.lower() in {"grand total", "total"}:
            continue
        names.append(raw)
        raw_rows.append(r)
    known = build_known_cities(names)
    out = []
    for r in raw_rows:
        raw = str(r[0]).strip()
        contact = str(r[1]).strip() if r[1] and str(r[1]).strip() not in {"", "None"} else ""
        mobile = r[2] if len(r) > 2 else None
        tel = r[3] if len(r) > 3 else None
        email = str(r[4]).strip() if len(r) > 4 and r[4] else ""
        biz0, city0, extra0 = split_name(raw, known)
        biz, city_from_biz, extra_from_biz = clean_business(biz0, city0, known)
        city, extra_from_city = clean_city(city0 or city_from_biz, known)
        extra = merge_extra(extra0, extra_from_biz, extra_from_city)
        if contact:
            extra = merge_extra(extra, f"contact: {contact}")
        p, s, notes = extract_phones(mobile, tel)
        if not p and not s:
            p, s, notes = extract_phones(tel, mobile)
        person = title_case(contact) if contact else None
        extra_bits = [extra] if extra else []
        if email:
            extra_bits.append(f"email: {email}")
        if notes:
            extra_bits.append(notes)
        additional = merge_extra(*extra_bits) or None
        out.append(
            {
                "original_name": raw,
                "business_name": (title_case(biz) if biz else "")[:500],
                "person_name": (person[:500] if person else None),
                "phone": digits_phone(p) or (p or None),
                "secondary_phone": digits_phone(s) or (s or None) or None,
                "city": city or None,
                "additional_details": additional,
                "email": email or None,
                "name_key": norm_key(biz or raw),
            }
        )
    wb.close()
    return out


def load_db():
    init_db()
    db = SessionLocal()
    try:
        cities = {c.id: c.name for c in db.query(City).all()}
        routes = {r.id: r.name for r in db.query(Route).all()}
        customers = (
            db.query(Customer)
            .filter(Customer.deleted_at.is_(None))
            .all()
        )
        rows = []
        phones: dict[str, int] = {}
        name_keys: dict[str, int] = {}
        for c in customers:
            rec = {
                "id": c.id,
                "business_name": c.business_name,
                "person_name": c.person_name,
                "phone": c.phone,
                "secondary_phone": c.secondary_phone,
                "city_name": cities.get(c.city_id) if c.city_id else None,
                "route_name": routes.get(c.route_id) if c.route_id else None,
                "additional_details": c.additional_details,
                "is_active": c.is_active,
            }
            rows.append(rec)
            for ph in (c.phone, c.secondary_phone):
                d = digits_phone(ph)
                if d:
                    phones.setdefault(d, c.id)
            name_keys.setdefault(norm_key(c.business_name or ""), c.id)
        return rows, phones, name_keys
    finally:
        db.close()


def main() -> None:
    sheet = load_sheet()
    db_rows, db_phones, db_names = load_db()
    missing = []
    matched_phone = 0
    matched_name = 0
    no_phone_and_missing = 0
    for rec in sheet:
        hit = None
        reason = None
        for ph in (digits_phone(rec["phone"]), digits_phone(rec["secondary_phone"])):
            if ph and ph in db_phones:
                hit = "phone"
                break
        if hit:
            matched_phone += 1
            continue
        if rec["name_key"] and rec["name_key"] in db_names:
            matched_name += 1
            continue
        if not rec["phone"]:
            no_phone_and_missing += 1
        missing.append(rec)

    summary = {
        "sheet_count": len(sheet),
        "db_count": len(db_rows),
        "matched_phone": matched_phone,
        "matched_name_only": matched_name,
        "missing": len(missing),
        "missing_without_phone": no_phone_and_missing,
    }
    print(json.dumps(summary, indent=2))

    payload = {"summary": summary, "missing": missing}
    OUT_JSON.write_text(json.dumps(payload, indent=2, ensure_ascii=False))

    wb = Workbook()
    ws = wb.active
    ws.title = "Missing from DB"
    headers = [
        "business_name",
        "person_name",
        "phone",
        "secondary_phone",
        "city",
        "additional_details",
        "original_name",
    ]
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center")
    for rec in missing:
        ws.append([rec.get(h) for h in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col, w in zip("ABCDEFG", (36, 22, 14, 14, 22, 40, 42)):
        ws.column_dimensions[col].width = w
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_JSON}")


if __name__ == "__main__":
    main()
