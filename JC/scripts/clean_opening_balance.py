#!/usr/bin/env python3
"""Clean DEBTOR WITH OPENING BALANCE → Cleaned + Needs Review sheets."""

from __future__ import annotations

import re
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "DEBTOR WITH OPENING BALANCE.XLSX"
OUT = ROOT / "debtors_opening_balance_cleaned.xlsx"

sys.path.insert(0, str(ROOT / "scripts"))
from clean_debtors import build_known_cities, split_name  # noqa: E402
from clean_debtors_r2 import (  # noqa: E402
    clean_business,
    clean_city,
    merge_extra,
    title_case,
)


def money100(v) -> str | None:
    if v is None or v == "":
        return None
    try:
        n = Decimal(str(v)) * Decimal("100")
    except (InvalidOperation, ValueError):
        return None
    # keep 2 decimals if needed, else int-like
    if n == n.to_integral_value():
        return format(n.quantize(Decimal("1")), "f")
    return format(n.quantize(Decimal("0.01")), "f")


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "A": 44, "B": 28, "C": 22, "D": 28,
        "E": 14, "F": 14, "G": 22,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def review_reasons(biz, city, extra, debit, credit) -> list[str]:
    rs = []
    if not biz:
        rs.append("no_business")
    if not city:
        rs.append("no_city")
    if extra:
        rs.append("has_extra")
    if debit is None and credit is None:
        rs.append("no_amount")
    return rs


def main() -> None:
    wb = load_workbook(SRC, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # data rows: after header row with Bills/Receipt/Debit/Credit
    start = 0
    for i, r in enumerate(rows):
        vals = [str(x).lower() if x is not None else "" for x in r]
        if "debit" in vals and "credit" in vals:
            start = i + 1
            break

    raw_names = []
    records = []
    for r in rows[start:]:
        name = r[0]
        if name is None or not str(name).strip():
            continue
        label = str(name).strip()
        if label.lower() in {"grand total", "total"}:
            continue
        if label == "0":
            continue
        # cols: Particulars, Bills, Receipt, Debit, Credit
        debit, credit = r[3], r[4]
        raw_names.append(label)
        records.append((label, debit, credit))

    known = build_known_cities(raw_names)

    out = Workbook()
    cleaned = out.active
    cleaned.title = "Cleaned"
    cols = [
        "original_name", "business_name", "city", "extra_details",
        "debit", "credit",
    ]
    cleaned.append(cols)
    review = out.create_sheet("Needs Review")
    review.append(cols + ["review_reason"])

    n_c = n_r = 0
    for raw, debit, credit in records:
        biz, city, extra = split_name(raw, known)
        city2, city_extra = clean_city(city, known)
        biz2, maybe_city, biz_extra = clean_business(biz, city2, known)
        if maybe_city and not city2:
            city2 = maybe_city
        # trailing junk on raw often left in business (IMLI after cash)
        extra2 = merge_extra(extra, city_extra, biz_extra)
        # move leftover ops tokens still stuck on city/biz
        if re.search(r"\b(imli|cash|block)\b", str(biz2 or ""), re.I):
            for m in re.finditer(r"\b(imli|cash|block)\b", str(biz2), re.I):
                extra2 = merge_extra(extra2, m.group(0))
            biz2 = re.sub(r"\b(imli|cash|block)\b", " ", str(biz2), flags=re.I)
            biz2 = title_case(re.sub(r"\s+", " ", biz2).strip())

        d100 = money100(debit)
        c100 = money100(credit)
        row = [raw, biz2 or None, city2 or None, extra2 or None, d100, c100]
        rs = review_reasons(biz2, city2, extra2, d100, c100)
        if rs:
            review.append(row + [", ".join(rs)])
            n_r += 1
        else:
            cleaned.append(row)
            n_c += 1

    style_header(cleaned)
    style_header(review)
    review.column_dimensions["G"].width = 28

    out.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Cleaned (sure): {n_c}")
    print(f"Needs Review: {n_r}")
    print(f"Total parties: {n_c + n_r}")


if __name__ == "__main__":
    main()
