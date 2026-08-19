#!/usr/bin/env python3
"""Import cleaned ALL ITEM STOCK rows as year_group=2015-25 products + opening stock."""

from __future__ import annotations

import os
import re
import sys
from collections import Counter
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "backend"
SRC = ROOT / "ALL ITEM STOCK (30-JUN-26).xlsx"
OUT = ROOT / "stock_products_cleaned.xlsx"
YEAR = "2015-25"
DEFAULT_VENDOR_NAME = "LEGACY STOCK"

sys.path.insert(0, str(BACKEND))
os.chdir(BACKEND)

from app.db.session import SessionLocal, init_db  # noqa: E402
from app.models.catalog_lookup import CatalogLookup  # noqa: E402
from app.models.catalog_product import CatalogProduct  # noqa: E402
from app.models.stock import StockBalance  # noqa: E402
from app.models.city import City  # noqa: E402
from app.models.vendor import Vendor  # noqa: E402
from app.services.pricing import coerce_selling_price  # noqa: E402
from app.services.stock_receipt import add_stock  # noqa: E402

CAT_ALIASES = {
    "PATRIKA": "PATRIKA",
    "PATRIKA**": "PATRIKA",
    "PATRIKA☺": "PATRIKA",
    "PATRIKA**☺": "PATRIKA",
    "PATRIKA-BIG": "PATRIKA",
    "PATRIKA-RED": "PATRIKA",
    "PATRIKA-CREAM": "PATRIKA",
    "PATRIKA-B": "PATRIKA",
    "CARD": "CARD",
    "CARDS": "CARD",
    "CARD**": "CARD",
    "CARDAS": "CARD",
    "CARD/": "CARD",
    "FARMAN": "CARD",
    "ENVELOPE": "ENVELOPE",
    "OLD": "OLD",
}


def to_dec(v):
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def qty_int(v) -> int | None:
    d = to_dec(v)
    if d is None:
        return None
    return int(d.to_integral_value(rounding=ROUND_HALF_UP))


def clean_particulars(raw: str):
    s = re.sub(r"\s+", " ", (raw or "").strip())
    code = None
    rest = s
    m = re.match(r"^(\d{3,4})\b\s*(.*)$", s)
    if m:
        code, rest = m.group(1), m.group(2).strip()
    rest2 = re.sub(r"^[-–]\s*B\s+", "", rest, flags=re.I)
    rest2 = re.sub(r"^-B\s+", "", rest2, flags=re.I)

    if re.match(r"^NO\.?\s+", rest2, re.I):
        category = "CARD"
        name = re.sub(r"^NO\.?\s+", "", rest2, flags=re.I).strip()
        return code, category, name or rest2, "ok"

    tokens = rest2.split()
    category = None
    name = rest2
    if tokens:
        first = tokens[0]
        first_key = re.split(r"[\(\-]", first)[0]
        key = first_key.upper()
        if key.startswith("PATRIKA"):
            category = "PATRIKA"
            name = re.sub(r"^PATRIKA[\w\*☺\-]*\s*", "", rest2, flags=re.I).strip() or rest2
        elif key in CAT_ALIASES:
            category = CAT_ALIASES[key]
            name = rest2[len(first) :].strip() or rest2
        else:
            category = None
            name = rest2

    name = re.sub(r"\{\{\s*\d+\s*\}\}", "", name).strip()
    name = re.sub(r"\s+", " ", name).strip(" -") or rest2 or s
    status = "ok"
    if not code:
        status = "needs_review_no_code"
    elif not category:
        status = "needs_review_no_category"
    return code, category, name, status


def ensure_lookup(db, lookup_type: str, value: str) -> None:
    val = value.strip()
    exists = (
        db.query(CatalogLookup)
        .filter(
            CatalogLookup.lookup_type == lookup_type,
            CatalogLookup.value == val,
            CatalogLookup.is_active.is_(True),
        )
        .one_or_none()
    )
    if exists:
        return
    db.add(CatalogLookup(lookup_type=lookup_type, value=val))
    db.flush()


def ensure_vendor(db) -> Vendor:
    row = (
        db.query(Vendor)
        .filter(Vendor.deleted_at.is_(None), Vendor.business_name == DEFAULT_VENDOR_NAME)
        .one_or_none()
    )
    if row:
        if not row.is_active:
            row.is_active = True
        return row
    city = (
        db.query(City)
        .filter(City.deleted_at.is_(None), City.name == "Legacy")
        .one_or_none()
    )
    if not city:
        city = City(name="Legacy", route_id=None, is_active=True)
        db.add(city)
        db.flush()
    # unique phone placeholder
    phone = "9999990001"
    while db.query(Vendor).filter(Vendor.phone == phone).one_or_none():
        phone = str(int(phone) + 1)
    row = Vendor(
        business_name=DEFAULT_VENDOR_NAME,
        phone=phone,
        city_id=city.id,
        is_active=True,
    )
    db.add(row)
    db.flush()
    return row


def set_absolute_stock(db, *, product: CatalogProduct, target: int) -> tuple[int, int]:
    bal = db.query(StockBalance).filter(StockBalance.catalog_product_id == product.id).first()
    current = int(bal.quantity_on_hand) if bal else 0
    delta = target - current
    if delta != 0:
        add_stock(
            db,
            catalog_product_id=product.id,
            our_product_id=product.our_product_id,
            quantity=delta,
            entry_type="opening_balance",
            reference_type="import",
            reference_id=0,
            party="legacy_stock_import",
            notes=f"Opening stock from ALL ITEM STOCK sheet → {target}",
        )
    return current, target


def load_excel_rows():
    wb = load_workbook(SRC, data_only=True)
    ws = wb.active
    rows = []
    for r in ws.iter_rows(min_row=7, values_only=True):
        if not r or not r[0]:
            continue
        particular = str(r[0]).strip()
        if particular.lower().startswith("grand"):
            continue
        code, cat, name, status = clean_particulars(particular)
        if not code:
            m = re.search(r"\b(RX-\d{2})\b", particular, re.I)
            if m:
                code = m.group(1).upper()
                status = "ok_soft_code"
                cat = cat or "ENVELOPE"
                name = name or particular
            elif re.search(r"\bgx\s*offset\b", particular, re.I):
                code = "GX OFFSET"
                status = "ok_soft_code"
                cat = cat or "ENVELOPE"
                name = name or particular
        rows.append(
            {
                "original": particular,
                "product_number": code,
                "category": cat,
                "name": name,
                "year_group": YEAR,
                "buying_price": to_dec(r[1]),
                "selling_price": to_dec(r[2]),
                "opening_qty": to_dec(r[3]),
                "status": status,
            }
        )
    wb.close()
    codes = [r["product_number"] for r in rows if r["product_number"]]
    dups = {c for c, n in Counter(codes).items() if n > 1}
    for r in rows:
        if r["product_number"] in dups and r["status"].startswith("ok"):
            r["status"] = "needs_review_duplicate_code"
    return rows


def write_report(rows, import_log):
    out = Workbook()
    ws1 = out.active
    ws1.title = "Cleaned"
    cols = [
        "status",
        "product_number",
        "category",
        "name",
        "year_group",
        "buying_price",
        "selling_price",
        "opening_qty",
        "original",
        "import_result",
        "db_product_id",
    ]
    ws1.append(cols)
    by_key = {(x.get("product_number"), x.get("original")): x for x in import_log}
    for r in rows:
        info = by_key.get((r["product_number"], r["original"])) or {}
        ws1.append(
            [
                r["status"],
                r["product_number"],
                r["category"],
                r["name"],
                r["year_group"],
                r["buying_price"],
                r["selling_price"],
                r["opening_qty"],
                r["original"],
                info.get("import_result"),
                info.get("db_product_id"),
            ]
        )

    ws2 = out.create_sheet("Needs Review")
    ws2.append(cols)
    for r in rows:
        if not str(r["status"]).startswith("ok"):
            info = by_key.get((r["product_number"], r["original"])) or {}
            ws2.append(
                [
                    r["status"],
                    r["product_number"],
                    r["category"],
                    r["name"],
                    r["year_group"],
                    r["buying_price"],
                    r["selling_price"],
                    r["opening_qty"],
                    r["original"],
                    info.get("import_result"),
                    info.get("db_product_id"),
                ]
            )

    ws3 = out.create_sheet("Imported")
    ws3.append(["import_result", "product_number", "db_product_id", "qty_set", "buy", "sell", "original"])
    for x in import_log:
        if str(x.get("import_result", "")).startswith(("created", "updated", "stock_set")):
            ws3.append(
                [
                    x.get("import_result"),
                    x.get("product_number"),
                    x.get("db_product_id"),
                    x.get("qty_set"),
                    x.get("buying_price"),
                    x.get("selling_price"),
                    x.get("original"),
                ]
            )

    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for ws in (ws1, ws2, ws3):
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
        ws.freeze_panes = "A2"
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions
    for i, w in enumerate([22, 14, 12, 40, 12, 12, 12, 12, 45, 18, 12], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
        ws2.column_dimensions[get_column_letter(i)].width = w
    out.save(OUT)


def main() -> None:
    print("Init DB…")
    init_db()
    rows = load_excel_rows()
    print(f"Excel rows: {len(rows)}")

    db = SessionLocal()
    import_log = []
    try:
        ensure_lookup(db, "year_group", YEAR)
        ensure_lookup(db, "year_group", "2026-27")
        ensure_lookup(db, "year_group", "2025-26")
        for cat in sorted({r["category"] for r in rows if r.get("category")}):
            ensure_lookup(db, "category", cat)
        vendor = ensure_vendor(db)
        db.commit()

        # index existing by our_product_id (any year) — avoid dual-year same number
        existing = (
            db.query(CatalogProduct)
            .filter(CatalogProduct.deleted_at.is_(None), CatalogProduct.is_active.is_(True))
            .all()
        )
        by_sku = {}
        for p in existing:
            by_sku.setdefault(str(p.our_product_id).strip().upper(), p)

        created = updated = stocked = skipped = 0
        for r in rows:
            if not str(r["status"]).startswith("ok"):
                import_log.append({**r, "import_result": "skipped_needs_review", "db_product_id": None})
                skipped += 1
                continue
            code = str(r["product_number"]).strip()
            buy = r["buying_price"] if r["buying_price"] is not None else Decimal("0.00")
            sell = coerce_selling_price(buy, r["selling_price"])
            qty = qty_int(r["opening_qty"])
            if qty is None:
                qty = 0

            key = code.upper()
            prod = by_sku.get(key)
            if prod:
                # update prices + put on YEAR if empty/other? keep existing year; set stock
                prod.buying_price = buy
                prod.selling_price = sell
                if not prod.category and r["category"]:
                    prod.category = r["category"]
                # if product has no year or was placeholder, set to 2015-25 only when currently 2025-26 matching sheet?
                # User wants this sheet in 2015-25. If SKU already exists another year, just stock that row (no second SKU).
                db.flush()
                before, after = set_absolute_stock(db, product=prod, target=qty)
                updated += 1
                stocked += 1
                import_log.append(
                    {
                        **r,
                        "import_result": f"updated_existing_stock {before}->{after}",
                        "db_product_id": prod.id,
                        "qty_set": after,
                    }
                )
            else:
                prod = CatalogProduct(
                    our_product_id=code,
                    vendor_id=vendor.id,
                    vendor_product_id=(r["name"] or code)[:255],
                    category=r["category"],
                    series=None,
                    unit="pcs",
                    year_group=YEAR,
                    buying_price=buy,
                    selling_price=sell,
                    image_keys=[],
                    is_active=True,
                )
                db.add(prod)
                db.flush()
                by_sku[key] = prod
                before, after = set_absolute_stock(db, product=prod, target=qty)
                created += 1
                stocked += 1
                import_log.append(
                    {
                        **r,
                        "import_result": f"created_stock {after}",
                        "db_product_id": prod.id,
                        "qty_set": after,
                    }
                )

            if (created + updated) % 100 == 0:
                db.commit()
                print(f"  progress created={created} updated={updated}")

        db.commit()
        print(f"Created: {created}")
        print(f"Updated existing: {updated}")
        print(f"Stock set: {stocked}")
        print(f"Skipped needs review: {skipped}")
        write_report(rows, import_log)
        print(f"Wrote {OUT}")
        n = db.query(CatalogProduct).filter(CatalogProduct.deleted_at.is_(None), CatalogProduct.is_active.is_(True)).count()
        print(f"Active products now: {n}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
