#!/usr/bin/env python3
"""Compare MAIN DEBTORS BACK 3 YEAR sheet vs DB opening + outstanding."""
from __future__ import annotations

import json
import sys
from collections import defaultdict
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))
sys.path.insert(0, str(ROOT / "backend"))

from clean_debtors import build_known_cities, split_name  # noqa: E402
from clean_debtors_r2 import clean_business, clean_city, title_case  # noqa: E402
from merge_debtors_final import norm_key  # noqa: E402

SHEET = ROOT / "MAIN DEBTORS BACK 3 YEAR.XLSX"
MISSING = ROOT / "debtors_missing_from_db.json"
OUT_JSON = ROOT / "debtors_opening_compare.json"
OUT_XLSX = ROOT / "debtors_opening_compare.xlsx"

TOL = Decimal("1.00")  # rupees


def money(v) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        d = Decimal(str(v).replace(",", "").strip())
    except Exception:
        return None
    if d == 0:
        return Decimal("0")
    return d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def parse_sheet():
    wb = load_workbook(SHEET, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    start = 10
    for i, r in enumerate(rows):
        vals = [str(c).strip().lower() if c is not None else "" for c in r]
        if "particulars" in vals and "bills" in vals:
            start = i + 1
            break
    raw_names = []
    recs = []
    for r in rows[start:]:
        name = r[0]
        if name is None or not str(name).strip():
            continue
        label = str(name).strip()
        if label.lower() in {"grand total", "total"}:
            continue
        raw_names.append(label)
        recs.append((label, money(r[1]), money(r[2]), money(r[3])))
    known = build_known_cities(raw_names)
    out = []
    for raw, bills, receipt, close in recs:
        biz, city, extra = split_name(raw, known)
        city2, city_extra = clean_city(city, known)
        biz2, maybe_city, biz_extra = clean_business(biz, city2, known)
        if maybe_city and not city2:
            city2 = maybe_city
        extra_parts = [p for p in [extra, city_extra, biz_extra] if p]
        extra2 = " | ".join(dict.fromkeys(extra_parts)) or None
        biz_name = title_case(biz2) if biz2 else None
        out.append({
            "original_name": raw,
            "business_name": biz_name,
            "city": title_case(city2) if city2 else None,
            "extra_details": extra2,
            "name_key": norm_key(biz2 or raw),
            "raw_key": norm_key(raw),
            "bills": str(bills) if bills is not None else None,
            "receipt": str(receipt) if receipt is not None else None,
            "closing": str(close) if close is not None else None,
        })
    return out


def load_db():
    import os
    from dotenv import load_dotenv
    load_dotenv(ROOT / "backend" / ".env")
    os.chdir(ROOT / "backend")
    from sqlalchemy import case, func

    from app.db.session import SessionLocal, init_db
    from app.models.accounts_receivable import ArLedgerEntry
    from app.models.city import City
    from app.models.customer import Customer

    init_db()
    db = SessionLocal()
    try:
        cities = {c.id: c.name for c in db.query(City).all()}
        customers = (
            db.query(Customer)
            .filter(Customer.deleted_at.is_(None))
            .order_by(Customer.id)
            .all()
        )
        ledger = {
            cid: (Decimal(str(ob or 0)).quantize(Decimal("0.01")),
                  Decimal(str(outst or 0)).quantize(Decimal("0.01")))
            for cid, ob, outst in db.query(
                ArLedgerEntry.customer_id,
                func.coalesce(
                    func.sum(case((ArLedgerEntry.entry_type == "opening_balance", ArLedgerEntry.amount), else_=0)),
                    0,
                ),
                func.coalesce(func.sum(ArLedgerEntry.amount), 0),
            ).group_by(ArLedgerEntry.customer_id)
        }
        rows = []
        for c in customers:
            ob, outst = ledger.get(c.id, (Decimal("0.00"), Decimal("0.00")))
            rows.append({
                "id": c.id,
                "business_name": c.business_name,
                "person_name": c.person_name,
                "phone": c.phone,
                "secondary_phone": c.secondary_phone,
                "city": cities.get(c.city_id),
                "name_key": norm_key(c.business_name or ""),
                "opening": str(ob),
                "outstanding": str(outst),
            })
        return rows
    finally:
        db.close()


def d(v) -> Decimal:
    if v is None or v == "":
        return Decimal("0")
    return Decimal(str(v))


def main():
    sheet = parse_sheet()
    missing = json.loads(MISSING.read_text())["missing"]
    db_rows = load_db()

    by_raw: dict[str, list] = defaultdict(list)
    by_key: dict[str, list] = defaultdict(list)
    for r in sheet:
        by_raw[norm_key(r["original_name"])].append(r)
        by_key[r["name_key"]].append(r)

    def pick_hit(m):
        hits = by_raw.get(norm_key(m.get("original_name") or "")) or []
        if not hits:
            hits = by_key.get(m.get("name_key") or "") or []
        if not hits:
            return None
        return max(hits, key=lambda h: d(h["closing"]))

    # --- missing 323 ---
    missing_out = []
    miss_hit = miss_none = 0
    for m in missing:
        hit = pick_hit(m)
        if not hit:
            miss_none += 1
            missing_out.append({**m, "sheet_found": False, "bills": None, "receipt": None, "closing": None})
            continue
        miss_hit += 1
        missing_out.append({
            **m,
            "sheet_found": True,
            "sheet_name": hit["original_name"],
            "bills": hit["bills"],
            "receipt": hit["receipt"],
            "closing": hit["closing"],
        })

    # Phone join via ONLY DEBTOR NEW (3-year sheet has no mobiles)
    from compare_debtors_to_db import load_sheet as load_phone_sheet

    phone_sheet = load_phone_sheet()
    phone_by_raw = {}
    phone_by_key = defaultdict(list)
    for rec in phone_sheet:
        phone_by_raw.setdefault(norm_key(rec["original_name"]), rec)
        if rec.get("name_key"):
            phone_by_key[rec["name_key"]].append(rec)

    db_by_phone: dict[str, dict] = {}
    db_by_key: dict[str, list] = defaultdict(list)
    for r in db_rows:
        db_by_key[r["name_key"]].append(r)
        for ph in (r.get("phone"), r.get("secondary_phone")):
            if ph:
                db_by_phone.setdefault(str(ph), r)

    used_db = set()
    used_sheet = set()
    pairs = []
    pair_how = {"phone": 0, "name": 0}
    for s in sheet:
        dbc = None
        how = None
        ph_rec = phone_by_raw.get(norm_key(s["original_name"]))
        if not ph_rec and s["name_key"] in phone_by_key:
            ph_rec = phone_by_key[s["name_key"]][0]
        if ph_rec:
            for ph in (ph_rec.get("phone"), ph_rec.get("secondary_phone")):
                if ph and ph in db_by_phone and db_by_phone[ph]["id"] not in used_db:
                    dbc = db_by_phone[ph]
                    how = "phone"
                    break
        if dbc is None:
            cands = [c for c in db_by_key.get(s["name_key"]) or [] if c["id"] not in used_db]
            if cands:
                dbc = cands[0]
                how = "name"
        if dbc is None:
            continue
        used_db.add(dbc["id"])
        used_sheet.add(id(s))
        pair_how[how] += 1
        pairs.append((s, dbc))

    # scale probe: among pairs with both > 0, which multiplier fits opening / outstanding
    def fit(pairs, db_field, mult):
        ok = near = far = 0
        diffs = []
        for s, dbc in pairs:
            sheet_amt = d(s["closing"]) * Decimal(mult)
            db_amt = d(dbc[db_field])
            if sheet_amt == 0 and db_amt == 0:
                ok += 1
                continue
            if sheet_amt == 0 or db_amt == 0:
                far += 1
                continue
            diff = abs(sheet_amt - db_amt)
            diffs.append(float(diff))
            if diff <= TOL:
                ok += 1
            elif diff <= Decimal("100"):
                near += 1
            else:
                far += 1
        return {"ok": ok, "near": near, "far": far, "n": len(pairs)}

    scales = {}
    for field in ("opening", "outstanding"):
        scales[field] = {m: fit(pairs, field, m) for m in (1, 100, 1000)}

    # pick best scale for outstanding first (sheet is a closing figure)
    best_mult = 1
    best_score = -1
    for m, st in scales["outstanding"].items():
        score = st["ok"] * 10 + st["near"] - st["far"]
        if score > best_score:
            best_score = score
            best_mult = m

    db_cmp = []
    match = mismatch = sheet_zero_db_pos = db_zero_sheet_pos = 0
    for s, dbc in pairs:
        sheet_rs = (d(s["closing"]) * Decimal(best_mult)).quantize(Decimal("0.01"))
        ob = d(dbc["opening"])
        outst = d(dbc["outstanding"])
        diff_ob = (sheet_rs - ob).quantize(Decimal("0.01"))
        diff_out = (sheet_rs - outst).quantize(Decimal("0.01"))
        ob_ok = abs(diff_ob) <= TOL
        out_ok = abs(diff_out) <= TOL
        if out_ok:
            match += 1
            status = "outstanding_match"
        elif ob_ok:
            mismatch += 1
            status = "opening_match_outstanding_diff"
        elif sheet_rs == 0 and outst > 0:
            sheet_zero_db_pos += 1
            status = "sheet_zero_db_has_due"
        elif outst == 0 and sheet_rs > 0:
            db_zero_sheet_pos += 1
            status = "db_zero_sheet_has_due"
        else:
            mismatch += 1
            status = "mismatch"
        db_cmp.append({
            "id": dbc["id"],
            "db_name": dbc["business_name"],
            "db_phone": dbc["phone"],
            "db_city": dbc["city"],
            "sheet_name": s["original_name"],
            "sheet_bills": s["bills"],
            "sheet_receipt": s["receipt"],
            "sheet_closing_raw": s["closing"],
            "sheet_closing_inr": str(sheet_rs),
            "db_opening": str(ob),
            "db_outstanding": str(outst),
            "diff_vs_opening": str(diff_ob),
            "diff_vs_outstanding": str(diff_out),
            "status": status,
        })

    db_not_on_sheet = [r for r in db_rows if r["id"] not in used_db]
    sheet_not_in_db = [s for s in sheet if id(s) not in used_sheet]

    summary = {
        "sheet_file": SHEET.name,
        "sheet_period": "1-Jul-23 to 15-Aug-26",
        "sheet_note": "Tally group closing (Bills / Receipt / Closing Balance), not a classic opening dump.",
        "sheet_rows": len(sheet),
        "db_customers": len(db_rows),
        "missing_323": len(missing),
        "scale_probe": scales,
        "chosen_inr_multiplier": best_mult,
        "missing": {
            "found_on_sheet": miss_hit,
            "not_on_sheet": miss_none,
            "found_with_closing_gt_0": sum(1 for r in missing_out if r.get("sheet_found") and d(r.get("closing")) > 0),
        },
        "db_compare": {
            "matched_by_name": len(pairs),
            "db_not_on_sheet": len(db_not_on_sheet),
            "sheet_not_in_db": len(sheet_not_in_db),
            "outstanding_match": match,
            "mismatch": mismatch,
            "sheet_zero_db_has_due": sheet_zero_db_pos,
            "db_zero_sheet_has_due": db_zero_sheet_pos,
        },
        "sheet_closing_raw_total": str(sum((d(s["closing"]) for s in sheet), Decimal("0"))),
        "db_opening_total": str(sum((d(r["opening"]) for r in db_rows), Decimal("0"))),
        "db_outstanding_total": str(sum((d(r["outstanding"]) for r in db_rows), Decimal("0"))),
    }

    OUT_JSON.write_text(json.dumps({
        "summary": summary,
        "missing": missing_out,
        "db_compare": db_cmp,
        "db_not_on_sheet": [
            {"id": r["id"], "name": r["business_name"], "phone": r["phone"],
             "opening": r["opening"], "outstanding": r["outstanding"]}
            for r in db_not_on_sheet
        ],
        "sheet_not_in_db": sheet_not_in_db,
    }, indent=2, ensure_ascii=False))

    wb = Workbook()
    s1 = wb.active
    s1.title = "Missing 323"
    s1.append(["business_name", "person_name", "phone", "city", "sheet_found",
               "sheet_name", "bills", "receipt", "closing_raw"])
    for r in missing_out:
        s1.append([
            r.get("business_name"), r.get("person_name"), r.get("phone"), r.get("city"),
            r.get("sheet_found"), r.get("sheet_name"), r.get("bills"), r.get("receipt"), r.get("closing"),
        ])
    s2 = wb.create_sheet("DB vs sheet")
    s2.append(["id", "db_name", "db_phone", "sheet_name", "sheet_closing_raw", "sheet_inr",
               "db_opening", "db_outstanding", "diff_opening", "diff_outstanding", "status"])
    for r in db_cmp:
        s2.append([
            r["id"], r["db_name"], r["db_phone"], r["sheet_name"], r["sheet_closing_raw"],
            r["sheet_closing_inr"], r["db_opening"], r["db_outstanding"],
            r["diff_vs_opening"], r["diff_vs_outstanding"], r["status"],
        ])
    s3 = wb.create_sheet("DB not on sheet")
    s3.append(["id", "name", "phone", "opening", "outstanding"])
    for r in db_not_on_sheet:
        s3.append([r["id"], r["business_name"], r["phone"], r["opening"], r["outstanding"]])
    wb.save(OUT_XLSX)
    print(json.dumps(summary, indent=2))
    print(f"wrote {OUT_JSON}")
    print(f"wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
