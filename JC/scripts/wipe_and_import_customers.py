#!/usr/bin/env python3
"""Wipe customer/order/stock data (keep catalog), import from debtors_merged_cleaned.xlsx."""

from __future__ import annotations

import os
import re
import sys
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import func, text

ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "backend"
MERGED = ROOT / "debtors_merged_cleaned.xlsx"
OUT = ROOT / "debtors_import_result.xlsx"
AS_ON = date(2025, 7, 1)

sys.path.insert(0, str(BACKEND))
os.chdir(BACKEND)

# load .env via pydantic settings
from app.db.session import SessionLocal, init_db, engine  # noqa: E402
from app.models.city import City  # noqa: E402
from app.models.customer import Customer  # noqa: E402
from app.services.ar_ledger import set_opening_balance  # noqa: E402
from app.services.passwords import hash_password  # noqa: E402


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def wipe(db) -> dict:
    """Hard-delete transactional data. Keep catalog/products/vendors/cities/routes/staff."""
    counts = {}
    stmts = [
        # customer side
        ("jc_customer_return_lines", "DELETE FROM jc_customer_return_lines"),
        ("jc_customer_returns", "DELETE FROM jc_customer_returns"),
        ("jc_ar_ledger_entries", "DELETE FROM jc_ar_ledger_entries"),
        ("jc_customer_bill_lines", "DELETE FROM jc_customer_bill_lines"),
        ("jc_customer_bills", "DELETE FROM jc_customer_bills"),
        ("jc_customer_order_lines", "DELETE FROM jc_customer_order_lines"),
        ("jc_customer_order_placements", "DELETE FROM jc_customer_order_placements"),
        ("jc_customer_open_lines", "DELETE FROM jc_customer_open_lines"),
        ("jc_customer_orders", "DELETE FROM jc_customer_orders"),
        ("jc_customer_ar_accounts", "DELETE FROM jc_customer_ar_accounts"),
        ("jc_customers", "DELETE FROM jc_customers"),
        # vendor orders
        ("jc_vendor_order_lines", "DELETE FROM jc_vendor_order_lines"),
        ("jc_vendor_order_placements", "DELETE FROM jc_vendor_order_placements"),
        ("jc_vendor_open_lines", "DELETE FROM jc_vendor_open_lines"),
        ("jc_vendor_orders", "DELETE FROM jc_vendor_orders"),
        # stock (keep catalog)
        ("jc_debit_notes", "DELETE FROM jc_debit_notes"),
        ("jc_stock_receipt_lines", "DELETE FROM jc_stock_receipt_lines"),
        ("jc_stock_receipts", "DELETE FROM jc_stock_receipts"),
        ("jc_stock_ledger", "DELETE FROM jc_stock_ledger"),
        ("jc_stock_balances", "DELETE FROM jc_stock_balances"),
        # AP tied to stock/payments — clear with stock wipe
        ("jc_ap_ledger_entries", "DELETE FROM jc_ap_ledger_entries"),
        ("jc_vendor_ap_accounts", "DELETE FROM jc_vendor_ap_accounts"),
    ]
    for name, sql in stmts:
        try:
            res = db.execute(text(sql))
            counts[name] = res.rowcount
        except Exception as e:
            counts[name] = f"err:{e}"
            db.rollback()
            raise
    db.commit()
    return counts


def load_merged_rows() -> list[dict]:
    wb = load_workbook(MERGED, data_only=True)
    rows_out = []
    for sn in ("Cleaned", "Needs Review"):
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        header = [str(h or "") for h in next(ws.iter_rows(max_row=1, values_only=True))]
        idx = {h: i for i, h in enumerate(header)}

        def g(r, name, default=None):
            i = idx.get(name)
            if i is None or i >= len(r):
                return default
            return r[i]

        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or not (g(r, "original_name") or g(r, "business_name")):
                continue
            rows_out.append({
                "original_name": g(r, "original_name"),
                "business_name": g(r, "business_name"),
                "city": g(r, "city"),
                "extra": g(r, "extra") or g(r, "extra_details"),
                "primary_phone": g(r, "primary_phone"),
                "secondary_phone": g(r, "secondary_phone"),
                "opening_balance": g(r, "opening_balance"),
                "person_hint": None,
            })
    wb.close()
    # extract contact: from extra into person_hint, strip from additional_details
    for row in rows_out:
        extra = str(row["extra"] or "")
        m = re.search(r"(?i)(?:^|\|\s*)contact:\s*([^|]+)", extra)
        if m:
            row["person_hint"] = m.group(1).strip()
            extra = re.sub(r"(?i)(?:^|\|\s*)contact:\s*[^|]+", "", extra)
            extra = re.sub(r"\s*\|\s*\|\s*", " | ", extra).strip(" |")
            row["extra"] = extra or None
        # drop email: bits from additional details? keep them — useful
    return rows_out


def normalize_phone(v) -> str | None:
    if v is None:
        return None
    d = re.sub(r"\D+", "", str(v))
    if len(d) == 11 and d.startswith("0"):
        d = d[1:]
    if len(d) == 12 and d.startswith("91"):
        d = d[2:]
    if len(d) == 10 and d[0] in "6789":
        return d
    return None


def parse_amount(v) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        n = Decimal(str(v))
    except (InvalidOperation, ValueError):
        return None
    if n <= 0:
        return None
    return n.quantize(Decimal("0.01"))


def get_or_create_city(db, city_name: str | None, cache: dict[str, int]) -> int | None:
    if not city_name or not str(city_name).strip():
        return None
    name = re.sub(r"\s+", " ", str(city_name).strip())
    key = name.lower()
    if key in cache:
        return cache[key]
    existing = (
        db.query(City)
        .filter(City.deleted_at.is_(None))
        .filter(func.lower(City.name) == key)
        .one_or_none()
    )
    if existing:
        if not existing.is_active:
            existing.is_active = True
            existing.deleted_at = None
        cache[key] = existing.id
        return existing.id
    row = City(name=name, route_id=None, is_active=True)
    db.add(row)
    db.flush()
    cache[key] = row.id
    return row.id


def main() -> None:
    print("Init DB / migrate…")
    init_db()
    db = SessionLocal()
    try:
        print("Wiping customers, orders, stock…")
        counts = wipe(db)
        for k, v in counts.items():
            print(f"  {k}: {v}")

        print("Loading merged Excel…")
        rows = load_merged_rows()
        print(f"  rows: {len(rows)}")

        city_cache: dict[str, int] = {}
        used_phones: set[str] = set()
        created = []
        not_created = []

        for row in rows:
            phone = normalize_phone(row["primary_phone"])
            secondary = normalize_phone(row["secondary_phone"])
            if secondary and secondary == phone:
                secondary = None
            biz = (row["business_name"] or "").strip()
            if not phone:
                not_created.append({**row, "reason": "no_phone"})
                continue
            if not biz:
                not_created.append({**row, "reason": "no_business_name"})
                continue
            if phone in used_phones:
                not_created.append({**row, "reason": "duplicate_phone_in_sheet"})
                continue
            # skip if somehow still in DB
            clash = db.query(Customer).filter(Customer.phone == phone).one_or_none()
            if clash:
                not_created.append({**row, "reason": f"phone_exists_id_{clash.id}"})
                continue

            city_id = get_or_create_city(db, row.get("city"), city_cache)
            route_id = None
            if city_id:
                city = db.get(City, city_id)
                route_id = city.route_id if city else None

            person = (row.get("person_hint") or "").strip() or None
            extra = (str(row.get("extra")).strip() if row.get("extra") else None) or None
            opening = parse_amount(row.get("opening_balance"))

            cust = Customer(
                business_name=biz[:500],
                person_name=person[:500] if person else None,
                phone=phone,
                password_hash=hash_password(phone[-4:]),
                secondary_phone=secondary,
                additional_details=extra,
                city_id=city_id,
                route_id=route_id,
                credit_limit=None,
                credit_override=False,
                is_active=True,
            )
            db.add(cust)
            db.flush()
            if opening is not None:
                set_opening_balance(
                    db,
                    customer_id=cust.id,
                    amount=opening,
                    as_on=AS_ON,
                    actor_type="admin",
                    actor_id=None,
                    actor_name="import",
                )
            used_phones.add(phone)
            created.append({
                **row,
                "customer_id": cust.id,
                "phone": phone,
                "secondary_phone": secondary,
                "opening_balance": str(opening) if opening is not None else None,
                "city_id": city_id,
                "person_name": person,
                "additional_details": extra,
            })

            if len(created) % 100 == 0:
                db.commit()
                print(f"  created {len(created)}…")

        db.commit()
        print(f"Created: {len(created)}")
        print(f"Not created: {len(not_created)}")

        # report workbook
        out = Workbook()
        ws_ok = out.active
        ws_ok.title = "Created in software"
        ok_cols = [
            "customer_id", "business_name", "person_name", "primary_phone", "secondary_phone",
            "city", "additional_details", "opening_balance", "original_name",
        ]
        ws_ok.append(ok_cols)
        for r in created:
            ws_ok.append([
                r["customer_id"], r["business_name"], r.get("person_name"),
                r["phone"], r.get("secondary_phone"), r.get("city"),
                r.get("additional_details"), r.get("opening_balance"), r.get("original_name"),
            ])

        ws_no = out.create_sheet("Not created")
        no_cols = [
            "reason", "business_name", "primary_phone", "secondary_phone",
            "city", "extra", "opening_balance", "original_name",
        ]
        ws_no.append(no_cols)
        for r in not_created:
            ws_no.append([
                r.get("reason"), r.get("business_name"), r.get("primary_phone"),
                r.get("secondary_phone"), r.get("city"), r.get("extra"),
                r.get("opening_balance"), r.get("original_name"),
            ])

        style_header(ws_ok)
        style_header(ws_no)
        for col, w in zip("ABCDEFGHI", [12, 28, 18, 14, 14, 20, 28, 14, 40]):
            ws_ok.column_dimensions[col].width = w
            ws_no.column_dimensions[col].width = w

        out.save(OUT)
        print(f"Wrote {OUT}")

        # sanity
        n_cust = db.query(Customer).filter(Customer.deleted_at.is_(None), Customer.is_active.is_(True)).count()
        print(f"Active customers in DB now: {n_cust}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
