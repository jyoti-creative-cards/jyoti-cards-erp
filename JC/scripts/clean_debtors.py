#!/usr/bin/env python3
"""Clean Tally debtors Excel → business_name, city, extra_details, phones."""

from __future__ import annotations

import re
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "debtors .xlsx"
OUT = ROOT / "debtors_cleaned.xlsx"

EXTRA_EXACT = {
    "cash", "block", "bloc", "full", "full bill", "old", "new", "strict",
    "strct", "strtct", "stct", "billing", "retail", "jb", "imli", "su", "ak",
    "f", "h", "a", "w", "r", "g", "j", "g*", "j*", "g**", "j**", "1st", "first",
}

# Words that are almost never a city when trailing
NOT_CITY = {
    "press", "printers", "printer", "graphics", "cards", "card", "offset",
    "computer", "computers", "photocopy", "stationery", "traders", "trading",
    "suppliers", "studio", "collection", "advertising", "packaging", "paper",
    "bhai", "bhaiya", "ji", "show", "room", "show-room", "enterprise",
    "enterprises", "services", "sales", "mart", "centre", "center", "palace",
    "online", "creation", "innovations", "division", "kutir", "patrika",
}

STATE_WORDS = {
    "rajasthan", "raj", "rj", "mp", "madhya pradesh", "gujarat", "maharashtra",
    "indore",  # sometimes used as district qualifier in paren
}


def title_token(tok: str) -> str:
    if not tok:
        return tok
    if re.fullmatch(r"(?:[A-Za-z]\.){1,6}", tok) or re.fullmatch(r"[A-Za-z](?:\.[A-Za-z])+\.?", tok):
        return tok.upper()
    if tok.isupper() and len(tok) <= 3 and tok.isalpha():
        return tok.upper()
    if "-" in tok and not tok.startswith("-"):
        return "-".join(title_token(p) if p else p for p in tok.split("-"))
    if tok in {"&", "/", "-", "–"}:
        return tok
    return tok[:1].upper() + tok[1:].lower() if len(tok) > 1 else tok.upper()


def title_business(name: str) -> str:
    name = re.sub(r"\s+", " ", (name or "").strip())
    if not name:
        return ""
    parts = [title_token(tok) for tok in name.split(" ") if tok]
    out = " ".join(parts)
    out = re.sub(r"\s+,", ",", out)
    out = re.sub(r"\s+&", " &", out)
    return out.strip(" -–,")


def normalize_city(city: str) -> str:
    city = re.sub(r"\s+", " ", (city or "").strip(" -–,\t"))
    if not city:
        return ""
    # Title-ish, keep paren content
    def _fix(m: re.Match) -> str:
        inner = m.group(1).strip()
        return f"({title_business(inner)})"

    city = re.sub(r"\(([^)]*)\)", _fix, city)
    # title outside parens
    bits = []
    i = 0
    for m in re.finditer(r"\([^)]*\)", city):
        before = city[i:m.start()].strip()
        if before:
            bits.append(title_business(before))
        bits.append(m.group(0))
        i = m.end()
    tail = city[i:].strip()
    if tail:
        bits.append(title_business(tail))
    return " ".join(bits).strip()


PLACEISH = {
    "rajasthan", "dongargaon", "banswada", "jhalawad", "manawar", "dhar",
    "ratlam", "indore", "biora", "petlawad", "thandla", "jhabua", "mhow",
    "gandhwani", "manpur", "sonkach", "kannod", "alirajpur", "solapur",
    "barnagar", "tarana", "sanawad", "jn", "raj", "rj", "dewas", "ujjain",
    "khargone", "khandwa", "chirakhan", "shinghana", "singhana",
}


def is_extra_tag(tag: str) -> bool:
    t = re.sub(r"\s+", " ", (tag or "").strip())
    if not t or t in {"", "-"}:
        return True
    cleaned = t.strip()
    key = cleaned.lower()
    key_nostar = re.sub(r"[*]+$", "", key).strip()
    if key in EXTRA_EXACT or key_nostar in EXTRA_EXACT:
        return True
    if re.fullmatch(r"\d{2}-\d{2}", key_nostar):
        return True
    if re.search(r"strict|block|full|cash|old|new|due|bill|strct|stct", key, re.I):
        return True
    if re.search(r"\d", cleaned) and re.search(r"/|due|bill", cleaned, re.I):
        return True
    if re.fullmatch(r"[A-Za-z][A-Za-z.'\s]{1,30}", cleaned):
        words = cleaned.split()
        low = [w.lower().strip("*") for w in words]
        if len(words) == 1 and low[0] in PLACEISH:
            return False
        if len(words) == 1 and low[0] in STATE_WORDS:
            return False
        if any(w in {"ji", "bhai", "bhaiya", "madam"} for w in low):
            return True
        if words[-1].lower() in {"nagar", "road", "bazar", "ganj", "pura", "wadi"}:
            return False
        if "jila" in low:
            return False
        # multi-word person/note (not place): "Santosh Ji", "Kishor Ji", "Shri Trading"
        if len(words) >= 2:
            return True
        # single bare name — only extra when clearly person/freight nickname style
        # (city decision handles unknown place names elsewhere)
        if len(words) == 1 and len(words[0]) >= 3:
            return False
    return False


def looks_like_city(token: str, known: set[str]) -> bool:
    t = re.sub(r"\s+", " ", (token or "").strip())
    if not t:
        return False
    t_clean = re.sub(r"[*]+$", "", t).strip()
    t_clean = re.sub(r"^/\s*\*?", "", t_clean).strip()
    key = t_clean.lower()
    parts = key.split()
    if not parts:
        return False
    if any(p in NOT_CITY for p in parts):
        return False
    if parts[-1] in {"bhai", "bhaiya", "ji"}:
        return False
    if key in {"retail", "billing", "block", "cash", "imli", "new", "old"}:
        return False
    if not re.fullmatch(r"[A-Za-z][A-Za-z.\s-]{1,40}", t_clean):
        return False
    if key in known or (parts[0] in known):
        return True
    for k in known:
        if k.startswith(key) or key.startswith(k):
            if abs(len(k) - len(key)) <= 2:
                return True
    # unknown single token ≥4 letters → city
    if len(parts) == 1 and len(parts[0]) >= 4:
        return True
    # two-word places only if neither word is business junk
    if len(parts) == 2 and all(len(p) >= 3 and p not in NOT_CITY for p in parts):
        return True
    return False


def extract_parens(text: str) -> tuple[str, list[str]]:
    tags = []
    def _grab(m: re.Match) -> str:
        tags.append(m.group(1).strip())
        return " "
    cleaned = re.sub(r"\(([^)]*)\)", _grab, text)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" -–,")
    return cleaned, tags


def split_name(raw: str, known_cities: set[str]) -> tuple[str, str, str]:
    raw = re.sub(r"\s+", " ", (raw or "").strip())
    raw = raw.replace("–", "-")
    # normalize weird trailing -()
    raw = re.sub(r"-\s*\(\)\s*$", "", raw).strip()

    business = raw
    city = ""
    extras: list[str] = []

    # Prefer " - " then last city-like dash then comma
    m = re.search(r"\s+-\s+(.+)$", raw)
    if m:
        business = raw[: m.start()].strip()
        rest = m.group(1).strip()
    else:
        # tight dash before city: "PRESS -SOYATKALA" or "DHARNI-(BURHANPUR)"
        m2 = re.search(r"-\s*(\(?[A-Za-z].*)$", raw)
        if m2 and re.search(r"[A-Za-z]", raw[: m2.start()]):
            left = raw[: m2.start()].strip()
            right = m2.group(1).strip(" -")
            # don't split hyphenated brand like Aal-Awadh (no space in left AND short)
            if (" " in left or len(left) > 10) and right:
                business, rest = left, right
            else:
                rest = ""
                business = raw
        else:
            rest = ""
    # "GURUKRIPA PRESS - (NISARPUR)" → rest is "(NISARPUR)" already
    # "PARIHAR-Deapalpur" short left — try known city on right of last hyphen
    if not rest and re.search(r"-[A-Za-z]", raw):
        left, right = raw.rsplit("-", 1)
        if looks_like_city(re.sub(r"[()]", "", right).strip(), known_cities):
            business, rest = left.strip(), right.strip()

    if not rest and "," in raw:
        # AGRAWAL CARD, JODHPUR
        left, right = raw.rsplit(",", 1)
        if looks_like_city(re.sub(r"\(.*?\)", "", right).strip(), known_cities) or "(" in right:
            business, rest = left.strip(), right.strip()

    if rest:
        rest_no, tags = extract_parens(rest)
        city = rest_no
        main_city = bool(rest_no.strip())
        for tag in tags:
            if not tag:
                continue
            key = re.sub(r"[*]+$", "", tag.lower()).strip()
            # operational noise → extra
            if key in EXTRA_EXACT or re.fullmatch(r"\d{2}-\d{2}", key) or re.search(
                r"strict|block|full|cash|old|new|due|bill|strct|stct", key, re.I
            ):
                extras.append(tag)
                continue
            # after a city name: only keep known place/locality/state in city col
            if main_city:
                if key in PLACEISH or key in STATE_WORDS or key in known_cities:
                    city = f"{city} ({tag})".strip()
                else:
                    extras.append(tag)  # Vishnu / Gurjar / contact
                continue
            if looks_like_city(tag, known_cities) or key in PLACEISH or key in STATE_WORDS:
                city = tag if not city else f"{city} ({tag})".strip()
                continue
            if not city:
                city = tag
            else:
                extras.append(tag)
        city_words = city.split()
        while city_words:
            w = city_words[-1].strip("*/")
            if w.lower() in EXTRA_EXACT or re.fullmatch(r"\d{2}-\d{2}", w):
                extras.append(city_words.pop())
                continue
            break
        city = " ".join(city_words)
        city = re.sub(r"\s*/\s*\*?$", "", city).strip()
        # rest was only a code like SU / CASH — recover city from business tail
        if (not city or city.lower() in EXTRA_EXACT) and business:
            if city and city.lower() in EXTRA_EXACT:
                extras.append(city)
                city = ""
            words = business.split()
            if len(words) >= 2:
                for n in (2, 1):
                    cand = " ".join(words[-n:])
                    if looks_like_city(cand, known_cities):
                        city = cand
                        business = " ".join(words[:-n])
                        break
    else:
        # no separator — try last token(s) as city; parens → city or extra
        base, tags = extract_parens(business)
        business = base
        for tag in tags:
            if not tag:
                continue
            key = re.sub(r"[*]+$", "", tag.lower()).strip()
            if key in EXTRA_EXACT or re.search(r"strict|block|full|cash|old|new|billing|retail", key, re.I):
                extras.append(tag)
            elif looks_like_city(tag, known_cities) or key in PLACEISH:
                city = tag if not city else f"{city} ({tag})"
            elif is_extra_tag(tag):
                extras.append(tag)
            else:
                # bare paren with no dash — likely city (Depalpur, Dhamnod)
                city = tag if not city else city
                if city != tag and tag not in extras:
                    extras.append(tag)
        words = business.split()
        if not city and len(words) >= 2:
            for n in (2, 1):
                cand = " ".join(words[-n:])
                if looks_like_city(cand, known_cities):
                    city = cand
                    business = " ".join(words[:-n])
                    break
        # paren-only city cases: Laxmi Printing Press (Depalpur)
        if not city:
            for tag in tags:
                if tag and not is_extra_tag(tag) and looks_like_city(tag, known_cities):
                    city = tag
                    break

    # also pull trailing ** notes from business
    if "**" in business:
        business = business.replace("**", "").strip()
        if "marked **" not in " ".join(extras).lower():
            extras.append("**")

    business = title_business(business)
    city = normalize_city(city)

    # dedupe extras preserve order
    seen = set()
    extra_out = []
    for e in extras:
        e2 = re.sub(r"\s+", " ", e).strip()
        if not e2:
            continue
        key = e2.lower()
        if key in seen:
            continue
        seen.add(key)
        extra_out.append(e2)

    return business, city, " | ".join(extra_out)


def _phones_from_text(text) -> tuple[list[str], list[str], list[str]]:
    blob = str(text or "")
    blob = blob.replace(",", " ").replace("/", " ").replace("|", " ")
    blob = re.sub(r"(?i)\b(mo\.?|mb\.?|no\.?|madam|mobile)\b", " ", blob)
    notes: list[str] = []
    if re.search(r"e\+\d+", blob, re.I):
        return [], [], ["bad_phone_scientific"]
    mobiles: list[str] = []
    others: list[str] = []
    for run in re.findall(r"\d[\d\-]{6,}\d", blob):
        d = re.sub(r"\D", "", run)
        if len(d) == 11 and d.startswith("0"):
            d = d[1:]
        if len(d) == 12 and d.startswith("91"):
            d = d[2:]
        if len(d) == 10 and d[0] in "6789":
            if d not in mobiles:
                mobiles.append(d)
        elif 7 <= len(d) <= 11:
            if d not in others:
                others.append(d)
                notes.append(f"non_mobile:{d}")
    for d in re.findall(r"[6-9]\d{9}", re.sub(r"\D", " ", blob)):
        if d not in mobiles:
            mobiles.append(d)
    return mobiles, others, notes


def extract_phones(primary, secondary) -> tuple[str, str, str]:
    """Return primary, secondary, phone_notes."""
    m1, o1, n1 = _phones_from_text(primary)
    m2, o2, n2 = _phones_from_text(secondary)
    mobiles: list[str] = []
    for d in m1 + m2:
        if d not in mobiles:
            mobiles.append(d)
    p = mobiles[0] if mobiles else ""
    s = mobiles[1] if len(mobiles) > 1 else ""
    notes = n1 + n2
    if not p and (o1 or o2):
        # keep a landline only if no mobile
        s = (o1 or o2)[0]
    elif not s and o1:
        notes.append(f"landline_ignored:{o1[0]}")
    return p, s, " | ".join(notes)


def build_known_cities(names: list[str]) -> set[str]:
    cands: Counter[str] = Counter()
    for raw in names:
        raw = str(raw)
        m = re.search(r"\s+-\s+([^-(]+?)(?:\s*\(|$)", raw)
        if m:
            cands[m.group(1).strip().lower()] += 1
        elif "," in raw:
            part = re.sub(r"\(.*?\)", "", raw.split(",")[-1]).strip().lower()
            if part:
                cands[part] += 1
    # keep anything seen ≥2 times, plus all once if len>=4
    known = {k for k, v in cands.items() if v >= 2 and len(k) >= 3}
    known |= {k for k, v in cands.items() if v >= 1 and len(k) >= 4}
    # seed common
    known |= {
        "indore", "dewas", "ratlam", "ujjain", "dhar", "manawar", "khargone",
        "khandwa", "bhopal", "jhabua", "alirajpur", "pithampur", "mhow",
        "raipur", "rajkot", "pune", "surat", "nagpur", "jodhpur", "kota",
        "betul", "sehore", "jabalpur", "badwani", "barwani", "sanawad",
        "thandla", "petlawad", "kukshi", "maheshwar", "ashta", "jaora",
        "nau", "rau", "dhule", "beawar", "aurangabad", "nanded", "morbi",
        "anand", "vadodara", "amreli", "jalgaon", "akola", "satna",
    }
    return known


def main() -> None:
    wb = load_workbook(SRC, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    # row0 title, row1 headers, row2 blank
    raw_names = []
    records = []
    for r in rows[3:]:
        name = r[0]
        if name is None or not str(name).strip():
            continue
        raw_names.append(str(name).strip())
        records.append((str(name).strip(), r[3], r[4]))

    known = build_known_cities(raw_names)

    out_wb = Workbook()
    out = out_wb.active
    out.title = "Debtors Cleaned"
    out.append([
        "original_name",
        "business_name",
        "city",
        "extra_details",
        "primary_phone",
        "secondary_phone",
        "phone_notes",
    ])

    for raw, p, s in records:
        biz, city, extra = split_name(raw, known)
        phone_p, phone_s, phone_notes = extract_phones(p, s)
        out.append([raw, biz, city, extra, phone_p, phone_s, phone_notes])

    # review sheet: odd rows
    rev = out_wb.create_sheet("Needs Review")
    rev.append(["original_name", "business_name", "city", "extra_details", "reason"])
    for row in out.iter_rows(min_row=2, values_only=True):
        raw, biz, city, extra, pp, sp, pn = row
        reasons = []
        if not city:
            reasons.append("no_city")
        if not pp:
            reasons.append("no_primary_phone")
        if not biz:
            reasons.append("no_business")
        if pn:
            reasons.append("phone_notes")
        if reasons:
            rev.append([raw, biz, city, extra, ", ".join(reasons)])

    out_wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Rows: {len(records)}")
    print(f"Review rows: {rev.max_row - 1}")


if __name__ == "__main__":
    main()
